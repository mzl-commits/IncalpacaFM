"""
Exportadores Excel para el módulo de Inventario/Movimientos.
Genera reportes profesionales con diseño institucional negro corporativo,
gráficos y tablas de resumen con información de Órdenes de Trabajo.
"""
import io
from collections import defaultdict
from pathlib import Path

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.inventario.models import Movimiento

# ─── Rutas al logo institucional ─────────────────────────────────────────────
_BASE = Path(__file__).resolve().parent

LOGO_PATHS = [
    _BASE / "logo_incalpaca.png",
    _BASE.parent.parent.parent / "frontend" / "public" / "logo-incalpaca.png",
    _BASE.parent / "workorders" / "logo_brand.png",
]


def _get_logo_path():
    for p in LOGO_PATHS:
        if p.exists():
            return p
    return None


def _insert_logo(ws, cell="A1", width=125, height=36):
    p = _get_logo_path()
    if not p:
        return
    try:
        img = OpenpyxlImage(str(p))
        img.width = width
        img.height = height
        ws.add_image(img, cell)
    except Exception:
        pass


# ─── Paleta corporativa (negro puro, a tono con el logo INCALPACA) ────────────
COLOR_HEADER = "000000"      # Negro puro
COLOR_HEADER_FONT = "FFFFFF"
COLOR_SUBHEADER = "1C1C1E"   # Negro zinc ligeramente suavizado
COLOR_ROW_ALT = "F7F7F8"     # Gris ultra-sutil
COLOR_ROW_BASE = "FFFFFF"
COLOR_BORDER = "E4E4E7"      # Borde gris claro
COLOR_TEXT = "09090B"        # Negro casi puro

# Badges de tipo de movimiento
TIPO_COLORS = {
    "entrada": ("065F46", "FFFFFF"),   # Verde bosque oscuro
    "salida":  ("1E3A8A", "FFFFFF"),   # Azul marino profundo
    "baja":    ("7F1D1D", "FFFFFF"),   # Rojo oscuro elegante
}

MESES_ES = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


# ─── Utilidades de estilo ─────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> Border:
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_font(size=11) -> Font:
    return Font(bold=True, color=COLOR_HEADER_FONT, size=size, name="Calibri")


def _body_font(bold=False, size=10, color=COLOR_TEXT) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _write_header_row(ws, row: int, columns: list, bg=COLOR_HEADER):
    fill = _fill(bg)
    font = _header_font(size=10.5)
    ws.row_dimensions[row].height = 27
    for col, text in enumerate(columns, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()


def _write_data_row(ws, row: int, values: list,
                    is_alt=False, alignments=None, cell_styles=None):
    bg = COLOR_ROW_ALT if is_alt else COLOR_ROW_BASE
    base_fill = _fill(bg)
    ws.row_dimensions[row].height = 21
    for col, value in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=value)
        c.border = _thin_border()
        if cell_styles and col in cell_styles:
            bg_c, fc_c, bold_c = cell_styles[col]
            c.fill = _fill(bg_c)
            c.font = _body_font(bold=bold_c, color=fc_c)
        else:
            c.fill = base_fill
            c.font = _body_font()
        align = (alignments or {}).get(col, "left")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)


def _set_col_widths(ws, widths: list):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ─── Resolución de Orden de Trabajo ──────────────────────────────────────────

def _obtener_ot(mov) -> tuple[str, str]:
    """
    Retorna (tiene_ot, codigo_ot) para un Movimiento.

    La relación es:
        Movimiento ← SolicitudMovimiento (FK movimiento) → work_order (FK WorkOrder)

    Si mov.solicitud_origen (reverse manager) tiene una SolicitudMovimiento con work_order,
    se usa ese. De lo contrario, se revisa referencia_externa como fallback.
    """
    # Intentar resolver via SolicitudMovimiento (relación correcta en el modelo)
    try:
        solicitud = mov.solicitud_origen.select_related("work_order").first()
        if solicitud and solicitud.work_order_id:
            code = solicitud.work_order.code or f"OT-{solicitud.work_order_id}"
            return "SÍ", code
    except Exception:
        pass

    # Fallback: referencia_externa con algún indicador de OT
    ref = (mov.referencia_externa or "").strip()
    if ref:
        return "SÍ", ref

    return "NO", "Sin OT"


# ─── Hoja 1: Top 15 Materiales + KPI Cards + BarChart ────────────────────────

def _hoja_top_materiales(wb, movimientos):
    """Resumen ejecutivo: KPI Cards y Top 15 materiales con gráfico de barras."""
    ws = wb.create_sheet("Top 15 Materiales")

    # Estadísticas por material
    stats: dict = {}
    for mov in movimientos:
        mid = mov.material_id
        if not mid:
            continue
        if mid not in stats:
            stats[mid] = {
                "codigo": mov.material.codigo if mov.material else str(mid),
                "nombre": mov.material.nombre if mov.material else "—",
                "total": 0, "entradas": 0, "salidas": 0, "bajas": 0,
                "ots": set(),
            }
        st = stats[mid]
        st["total"] += 1
        tipo = mov.tipo
        if tipo == "entrada":
            st["entradas"] += mov.cantidad
        elif tipo == "salida":
            st["salidas"] += mov.cantidad
        elif tipo == "baja":
            st["bajas"] += mov.cantidad

        # Resolución OT (solo leer el atributo prefetched, no hacer query aquí)
        ref = (mov.referencia_externa or "").strip()
        if ref:
            st["ots"].add(ref)

    sorted_mats = sorted(stats.values(), key=lambda x: x["total"], reverse=True)

    # ── Cabecera: Logo + Título ──────────────────────────────────────────────
    _insert_logo(ws, cell="A1", width=125, height=36)
    ws.merge_cells("C1:H1")
    tc = ws["C1"]
    tc.value = "INCALPACA TOPS S.A. — REPORTE GENERAL DE MOVIMIENTOS DE ALMACÉN"
    tc.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    tc.fill = _fill(COLOR_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── Fila 2: fecha del reporte ────────────────────────────────────────────
    ws.merge_cells("C2:H2")
    date_cell = ws["C2"]
    date_cell.value = f"Generado el {timezone.localdate().strftime('%d/%m/%Y')}   |   Total de registros: {len(movimientos)}"
    date_cell.font = Font(bold=False, size=9, color="6B7280", name="Calibri")
    date_cell.fill = _fill("18181B")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── KPI Cards (fila 4 y 5) ──────────────────────────────────────────────
    max_mat = sorted_mats[0] if sorted_mats else {"codigo": "—", "nombre": "Sin datos", "total": 0}
    min_mat = sorted_mats[-1] if sorted_mats else {"codigo": "—", "nombre": "Sin datos", "total": 0}

    for start_col, label, mat in [
        ("A", "MATERIAL CON MAYOR ROTACIÓN", max_mat),
        ("E", "MATERIAL CON MENOR ROTACIÓN", min_mat),
    ]:
        end_col = chr(ord(start_col) + 3)  # 4 columns wide
        ws.merge_cells(f"{start_col}4:{end_col}4")
        hdr = ws[f"{start_col}4"]
        hdr.value = label
        hdr.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        hdr.fill = _fill(COLOR_SUBHEADER)
        hdr.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{start_col}5:{end_col}5")
        val = ws[f"{start_col}5"]
        val.value = f"{mat['codigo']} — {mat['nombre']}   ({mat['total']} movimientos)"
        val.font = Font(bold=True, size=10, color=COLOR_TEXT, name="Calibri")
        val.fill = _fill(COLOR_ROW_ALT)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = _thin_border()

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 24

    # ── Tabla Top 15 (fila 7 en adelante) ───────────────────────────────────
    top_15 = sorted_mats[:15]
    cols = ["#", "Código", "Nombre del Material", "Total Movs.", "Entradas", "Salidas", "Bajas", "Órdenes de Trabajo"]
    widths = [6, 14, 36, 13, 13, 13, 13, 32]
    _write_header_row(ws, 7, cols)
    _set_col_widths(ws, widths)
    _freeze(ws, "A8")

    align = {1: "center", 2: "center", 3: "left", 4: "center",
             5: "center", 6: "center", 7: "center", 8: "left"}

    for i, mat in enumerate(top_15, 1):
        ots = ", ".join(sorted(mat["ots"])) if mat["ots"] else "Sin OT"
        _write_data_row(ws, 7 + i, [
            i, mat["codigo"], mat["nombre"],
            mat["total"], mat["entradas"], mat["salidas"], mat["bajas"],
            ots,
        ], is_alt=(i % 2 == 0), alignments=align)

    # ── Gráfico de Barras ────────────────────────────────────────────────────
    if top_15:
        chart = BarChart()
        chart.type = "col"
        chart.style = 2          # estilo limpio sin ruido
        chart.title = "Top 15 — Total de Movimientos por Material"
        chart.y_axis.title = "Total de Movimientos"
        chart.x_axis.title = None
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.delete = False
        chart.legend = None
        chart.width = 26
        chart.height = 15

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showPercent = False

        # min_row=7 es el encabezado, titles_from_data=True lo toma como título de serie
        data = Reference(ws, min_col=4, min_row=7, max_row=7 + len(top_15))
        cats = Reference(ws, min_col=3, min_row=8, max_row=7 + len(top_15))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J7")


# ─── Hoja 2: Detalle completo de movimientos ─────────────────────────────────

def _hoja_detalle_general(wb, movimientos):
    """Lista completa con nombre de material, tipo, y columnas explícitas de OT."""
    ws = wb.create_sheet("Detalle de Movimientos")
    cols = [
        "Fecha", "Hora", "Código", "Nombre del Material", "Tipo",
        "Cantidad", "Empaques", "Responsable",
        "¿Tiene OT?", "Orden de Trabajo", "Referencia / Observaciones"
    ]
    widths = [13, 9, 13, 34, 13, 10, 10, 24, 11, 24, 34]
    _write_header_row(ws, 1, cols)
    _set_col_widths(ws, widths)
    _freeze(ws)

    tipo_cell_styles = {
        tipo: (bg, fc, True) for tipo, (bg, fc) in TIPO_COLORS.items()
    }
    align = {
        1: "center", 2: "center", 3: "center", 4: "left", 5: "center",
        6: "center", 7: "center", 8: "left", 9: "center", 10: "left", 11: "left",
    }

    for r, mov in enumerate(movimientos, 2):
        fecha = mov.fecha.date() if hasattr(mov.fecha, "date") else mov.fecha
        hora = mov.fecha.strftime("%H:%M") if hasattr(mov.fecha, "strftime") else ""
        resp = (mov.responsable.get_full_name() or mov.responsable.username
                if mov.responsable else "N/A")
        tiene_ot, ot_code = _obtener_ot(mov)
        t_style = tipo_cell_styles.get(mov.tipo)

        _write_data_row(ws, r, [
            fecha.strftime("%d/%m/%Y"),
            hora,
            mov.material.codigo if mov.material else "—",
            mov.material.nombre if mov.material else "—",
            mov.get_tipo_display(),
            mov.cantidad,
            mov.cantidad_cajas if mov.cantidad_cajas is not None else "",
            resp,
            tiene_ot,
            ot_code,
            (mov.observaciones or mov.referencia_externa or "—"),
        ], is_alt=(r % 2 == 0), alignments=align,
           cell_styles={5: t_style} if t_style else None)

    ws.auto_filter.ref = f"A1:K{max(2, len(movimientos) + 1)}"


# ─── Construcción de frecuencias agrupadas ───────────────────────────────────

def _construir_frecuencia_general(movimientos):
    """
    Agrupa movimientos por (fecha, material), (año, mes, material) y (año, material).
    Solo acumula tipos válidos: entrada, salida, baja.
    Devuelve (por_dia, por_mes, por_anio).
    """
    _TIPOS_VALIDOS = {"entrada", "salida", "baja"}

    def _nueva_fila():
        return {"nombre": "", "codigo": "", "entrada": 0, "salida": 0, "baja": 0, "ots": set()}

    por_dia = defaultdict(_nueva_fila)
    por_mes = defaultdict(_nueva_fila)
    por_anio = defaultdict(_nueva_fila)

    for mov in movimientos:
        if mov.tipo not in _TIPOS_VALIDOS:
            continue  # ignorar tipos inesperados
        fecha = mov.fecha.date() if hasattr(mov.fecha, "date") else mov.fecha
        anio, mes = fecha.year, fecha.month
        mid = mov.material_id
        nombre = mov.material.nombre if mov.material else "—"
        codigo = mov.material.codigo if mov.material else "—"
        cantidad = mov.cantidad  # campo NOT NULL en el modelo

        key_d, key_m, key_a = (fecha, mid), (anio, mes, mid), (anio, mid)

        for d in [por_dia[key_d], por_mes[key_m], por_anio[key_a]]:
            d["nombre"] = nombre
            d["codigo"] = codigo
            d[mov.tipo] += cantidad

        # OT via referencia_externa (sin query extra, ya está cargado)
        ref = (mov.referencia_externa or "").strip()
        if ref:
            for d in [por_dia[key_d], por_mes[key_m], por_anio[key_a]]:
                d["ots"].add(ref)

    return por_dia, por_mes, por_anio


# ─── Hojas de frecuencia (Por Día / Por Mes / Por Año) ───────────────────────

def _hoja_por_dia(wb, por_dia):
    ws = wb.create_sheet("Por Día")
    cols = ["Fecha", "Código", "Material", "Entradas", "Salidas", "Bajas", "Total", "Órdenes de Trabajo"]
    widths = [13, 13, 36, 10, 10, 10, 10, 32]
    _write_header_row(ws, 1, cols)
    _set_col_widths(ws, widths)
    _freeze(ws)
    align = {1: "center", 2: "center", 3: "left",
             4: "center", 5: "center", 6: "center", 7: "center", 8: "left"}

    rows = sorted(por_dia.items(), key=lambda x: x[0][0])
    for r, ((fecha, _), d) in enumerate(rows, 2):
        total = d["entrada"] + d["salida"] + d["baja"]
        ots = ", ".join(sorted(d["ots"])) if d["ots"] else "Sin OT"
        _write_data_row(ws, r, [
            fecha.strftime("%d/%m/%Y"), d["codigo"], d["nombre"],
            d["entrada"] or "", d["salida"] or "", d["baja"] or "", total, ots,
        ], is_alt=(r % 2 == 0), alignments=align)
    ws.auto_filter.ref = f"A1:H{max(2, len(rows) + 1)}"


def _hoja_por_mes(wb, por_mes):
    ws = wb.create_sheet("Por Mes")
    cols = ["Año", "Mes", "Código", "Material", "Entradas", "Salidas", "Bajas", "Total", "Órdenes de Trabajo"]
    widths = [8, 8, 13, 36, 10, 10, 10, 10, 32]
    _write_header_row(ws, 1, cols)
    _set_col_widths(ws, widths)
    _freeze(ws)
    align = {1: "center", 2: "center", 3: "center", 4: "left",
             5: "center", 6: "center", 7: "center", 8: "center", 9: "left"}

    rows = sorted(por_mes.items(), key=lambda x: (x[0][0], x[0][1]))
    for r, ((anio, mes, _), d) in enumerate(rows, 2):
        total = d["entrada"] + d["salida"] + d["baja"]
        ots = ", ".join(sorted(d["ots"])) if d["ots"] else "Sin OT"
        _write_data_row(ws, r, [
            anio, MESES_ES[mes], d["codigo"], d["nombre"],
            d["entrada"] or "", d["salida"] or "", d["baja"] or "", total, ots,
        ], is_alt=(r % 2 == 0), alignments=align)
    ws.auto_filter.ref = f"A1:I{max(2, len(rows) + 1)}"


def _hoja_por_anio(wb, por_anio):
    ws = wb.create_sheet("Por Año")
    cols = ["Año", "Código", "Material", "Entradas", "Salidas", "Bajas", "Total", "Órdenes de Trabajo"]
    widths = [8, 13, 36, 10, 10, 10, 10, 32]
    _write_header_row(ws, 1, cols)
    _set_col_widths(ws, widths)
    _freeze(ws)
    align = {1: "center", 2: "center", 3: "left",
             4: "center", 5: "center", 6: "center", 7: "center", 8: "left"}

    rows = sorted(por_anio.items(), key=lambda x: x[0][0])
    for r, ((anio, _), d) in enumerate(rows, 2):
        total = d["entrada"] + d["salida"] + d["baja"]
        ots = ", ".join(sorted(d["ots"])) if d["ots"] else "Sin OT"
        _write_data_row(ws, r, [
            anio, d["codigo"], d["nombre"],
            d["entrada"] or "", d["salida"] or "", d["baja"] or "", total, ots,
        ], is_alt=(r % 2 == 0), alignments=align)
    ws.auto_filter.ref = f"A1:H{max(2, len(rows) + 1)}"


# ─── Hojas de historial individual por material ───────────────────────────────

def _hoja_historial_material(wb, movimientos, material):
    """Hoja principal: encabezado con logo, tarjeta de material, tabla de movimientos con OTs."""
    ws = wb.create_sheet("Historial")

    mat_nombre = material.nombre if material else "Material"
    mat_cod = material.codigo if material else "—"
    mat_stock = getattr(material, "cantidad_total", 0) if material else 0
    mat_unidad = getattr(material, "unidad_medida_nombre", "") if material else ""

    # ── Cabecera: Logo + Título ──────────────────────────────────────────────
    _insert_logo(ws, cell="A1", width=125, height=36)
    ws.merge_cells("C1:J1")
    tc = ws["C1"]
    tc.value = "INCALPACA TOPS S.A. — HISTORIAL DE MOVIMIENTOS POR MATERIAL"
    tc.font = Font(bold=True, size=12.5, color="FFFFFF", name="Calibri")
    tc.fill = _fill(COLOR_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── Fila 2: tarjeta de datos del material ────────────────────────────────
    ws.merge_cells("C2:J2")
    sub = ws["C2"]
    sub.value = f"{mat_nombre}   |   Cód: {mat_cod}   |   Stock actual: {mat_stock} {mat_unidad}"
    sub.font = Font(bold=True, size=10.5, color="FFFFFF", name="Calibri")
    sub.fill = _fill(COLOR_SUBHEADER)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    sub.border = _thin_border()
    ws.row_dimensions[2].height = 22

    # ── Fila 4: encabezados de la tabla ─────────────────────────────────────
    cols = [
        "Fecha", "Hora", "Tipo", "Cantidad", "Empaques",
        "Responsable", "¿Tiene OT?", "Orden de Trabajo", "Observaciones / Referencia"
    ]
    widths = [13, 9, 13, 10, 10, 24, 11, 24, 40]
    _write_header_row(ws, 4, cols)
    _set_col_widths(ws, widths)
    _freeze(ws, cell="A5")

    tipo_cell_styles = {tipo: (bg, fc, True) for tipo, (bg, fc) in TIPO_COLORS.items()}
    align = {
        1: "center", 2: "center", 3: "center", 4: "center", 5: "center",
        6: "left", 7: "center", 8: "left", 9: "left",
    }

    for r, mov in enumerate(movimientos, 5):
        fecha = mov.fecha.date() if hasattr(mov.fecha, "date") else mov.fecha
        hora = mov.fecha.strftime("%H:%M") if hasattr(mov.fecha, "strftime") else ""
        resp = (mov.responsable.get_full_name() or mov.responsable.username
                if mov.responsable else "N/A")
        tiene_ot, ot_code = _obtener_ot(mov)
        t_style = tipo_cell_styles.get(mov.tipo)
        obs = " | ".join(filter(None, [mov.observaciones, mov.referencia_externa])) or "—"

        _write_data_row(ws, r, [
            fecha.strftime("%d/%m/%Y"), hora,
            mov.get_tipo_display(),
            mov.cantidad,
            mov.cantidad_cajas if mov.cantidad_cajas is not None else "",
            resp,
            tiene_ot,
            ot_code,
            obs,
        ], is_alt=(r % 2 == 0), alignments=align,
           cell_styles={3: t_style} if t_style else None)

    ws.auto_filter.ref = f"A4:I{max(5, len(movimientos) + 4)}"


def _hoja_resumen_dia_material(wb, movimientos):
    ws = wb.create_sheet("Resumen por Día")
    cols = ["Fecha", "Entradas", "Salidas", "Bajas", "Total", "Órdenes de Trabajo"]
    widths = [14, 12, 12, 12, 12, 32]
    _write_header_row(ws, 1, cols)
    _set_col_widths(ws, widths)
    _freeze(ws)
    align = {1: "center", 2: "center", 3: "center", 4: "center", 5: "center", 6: "left"}

    por_dia = defaultdict(lambda: {"entrada": 0, "salida": 0, "baja": 0, "ots": set()})
    for mov in movimientos:
        if mov.tipo not in {"entrada", "salida", "baja"}:
            continue
        fecha = mov.fecha.date() if hasattr(mov.fecha, "date") else mov.fecha
        por_dia[fecha][mov.tipo] += mov.cantidad
        ref = (mov.referencia_externa or "").strip()
        if ref:
            por_dia[fecha]["ots"].add(ref)

    for r, (fecha, d) in enumerate(sorted(por_dia.items()), 2):
        total = d["entrada"] + d["salida"] + d["baja"]
        ots = ", ".join(sorted(d["ots"])) if d["ots"] else "Sin OT"
        _write_data_row(ws, r, [
            fecha.strftime("%d/%m/%Y"),
            d["entrada"] or "", d["salida"] or "", d["baja"] or "", total, ots,
        ], is_alt=(r % 2 == 0), alignments=align)
    ws.auto_filter.ref = f"A1:F{max(2, len(por_dia) + 1)}"


def _hoja_resumen_mes_material(wb, movimientos):
    ws = wb.create_sheet("Resumen por Mes")
    cols = ["Año", "Mes", "Entradas", "Salidas", "Bajas", "Total", "Órdenes de Trabajo"]
    widths = [8, 8, 12, 12, 12, 12, 32]
    _write_header_row(ws, 1, cols)
    _set_col_widths(ws, widths)
    _freeze(ws)
    align = {1: "center", 2: "center", 3: "center", 4: "center",
             5: "center", 6: "center", 7: "left"}

    por_mes = defaultdict(lambda: {"entrada": 0, "salida": 0, "baja": 0, "ots": set()})
    for mov in movimientos:
        if mov.tipo not in {"entrada", "salida", "baja"}:
            continue
        fecha = mov.fecha.date() if hasattr(mov.fecha, "date") else mov.fecha
        por_mes[(fecha.year, fecha.month)][mov.tipo] += mov.cantidad
        ref = (mov.referencia_externa or "").strip()
        if ref:
            por_mes[(fecha.year, fecha.month)]["ots"].add(ref)

    for r, ((anio, mes), d) in enumerate(sorted(por_mes.items()), 2):
        total = d["entrada"] + d["salida"] + d["baja"]
        ots = ", ".join(sorted(d["ots"])) if d["ots"] else "Sin OT"
        _write_data_row(ws, r, [
            anio, MESES_ES[mes],
            d["entrada"] or "", d["salida"] or "", d["baja"] or "", total, ots,
        ], is_alt=(r % 2 == 0), alignments=align)
    ws.auto_filter.ref = f"A1:G{max(2, len(por_mes) + 1)}"


# ─── Punto de entrada principal ───────────────────────────────────────────────

def generar_excel_movimientos(material_id=None) -> tuple:
    """
    Genera el Excel de movimientos con diseño corporativo negro y logo oficial.

    - Sin material_id → 5 hojas: Top 15 + Detalle completo + Por Día + Por Mes + Por Año
    - Con material_id → 3 hojas: Historial + Resumen por Día + Resumen por Mes

    Acepta material_id como PK (int/str) o código de material (str).
    Devuelve (BytesIO, filename).
    """
    from apps.catalogo.models import Material as MaterialModel

    wb = Workbook()
    wb.remove(wb.active)
    hoy = timezone.localdate()

    if material_id:
        material = None
        try:
            material = MaterialModel.objects.get(pk=int(material_id))
        except (ValueError, TypeError, MaterialModel.DoesNotExist):
            try:
                material = MaterialModel.objects.get(codigo__iexact=str(material_id).strip())
            except MaterialModel.DoesNotExist:
                pass

        target_id = material.id if material else material_id
        movimientos = list(
            Movimiento.objects
            .filter(material_id=target_id)
            .select_related("material", "responsable")
            .prefetch_related("solicitud_origen__work_order")
            .order_by("-fecha")
        )
        _hoja_historial_material(wb, movimientos, material)
        _hoja_resumen_dia_material(wb, movimientos)
        _hoja_resumen_mes_material(wb, movimientos)

        codigo = (material.codigo or str(material_id)) if material else str(material_id)
        nombre_safe = (
            (material.nombre[:20] if material else "material")
            .replace("/", "-").replace("\\", "-").replace(":", "")
        )
        filename = f"historial_{codigo}_{nombre_safe}_{hoy.isoformat()}.xlsx"

    else:
        movimientos = list(
            Movimiento.objects
            .select_related("material", "responsable")
            .prefetch_related("solicitud_origen__work_order")
            .order_by("-fecha")
        )
        por_dia, por_mes, por_anio = _construir_frecuencia_general(movimientos)

        _hoja_top_materiales(wb, movimientos)
        _hoja_detalle_general(wb, movimientos)
        _hoja_por_dia(wb, por_dia)
        _hoja_por_mes(wb, por_mes)
        _hoja_por_anio(wb, por_anio)

        filename = f"movimientos_almacen_{hoy.isoformat()}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename
