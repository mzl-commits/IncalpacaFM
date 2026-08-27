import io
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


TEMPLATE_PATH = Path(__file__).resolve().parent / "plantillas" / "Formato_Inspeccion.xlsx"

# Nombres de hoja y tipos de formato
HOJA_MANUALES = "Manuales"
HOJA_INALAMBRICAS = "Electricas Inalambricas"
HOJA_CON_CABLE = "Electricas con cable "
HOJA_EPP = "EPP"
HOJA_ESCALERAS = "Escaleras"
HOJA_CAIDAS = "Equipos contra caidas"

RESULTADO_COLS = {"apta": "C", "requiere_reparacion": "E", "fuera_servicio": "G"}
ACCION_COLS = {
    "continua_servicio": "A",
    "enviar_reparacion": "C",
    "retirar_servicio": "D",
    "dar_baja": "E",
    "reemplazar": "F",
}

CONFIG_INDIVIDUAL_COMUN = {
    "campos": {
        "codigo_herramienta": "C8",
        "proxima_inspeccion": "G8",
        "marca": "C9",
        "inspector": "C10",
        "fecha_inspeccion": "C11",
        "nombre_herramienta": "C12",
    },
}

CONFIG_HOJAS = {
    HOJA_MANUALES: {
        "tipo": "grupal",
        "criterio_data_start": 15,
        "num_criterios_nativos": 16,   # filas 15-30 en el template Manuales
        "campos": {
            "tipo_herramienta": "C8",
            "responsable": "G8",
            "cant_inspeccionada": "C9",
            "fecha_inspeccion": "G9",
            "cant_apta": "C10",
            "proxima_inspeccion": "G10",
            "cant_no_apta": "C11",
        },
        "resultado_aptas_row": 44,
        "resultado_observaciones_row": 45,
        "observaciones_generales": "A48",
    },
    HOJA_INALAMBRICAS: {
        "tipo": "individual",
        "criterio_data_start": 16,
        "num_criterios_nativos": 20,   # aprox filas 16-35 en Electricas Inalambricas
        "campos": CONFIG_INDIVIDUAL_COMUN["campos"],
        "resultado_row": 37,
        "accion_row": 41,
        "observaciones_generales": "A44",
    },
    HOJA_CON_CABLE: {
        "tipo": "individual",
        "criterio_data_start": 16,
        "num_criterios_nativos": 18,   # aprox filas 16-33 en Electricas con cable
        "campos": CONFIG_INDIVIDUAL_COMUN["campos"],
        "resultado_row": 35,
        "accion_row": 39,
        "observaciones_generales": "A42",
    },
}

def _normalizar(texto):
    """Quita tildes/mayusculas y corrige mojibake comun (utf-8 mal leido como latin-1)."""
    if not texto:
        return ""
    try:
        texto = texto.encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.lower()

def _detectar_hoja(plantilla_nombre):
    """Detecta el tipo de formato/hoja según el nombre de la plantilla."""
    nombre = _normalizar(plantilla_nombre)
    if "caida" in nombre or "caída" in nombre or "arnes" in nombre:
        return HOJA_CAIDAS
    if "escalera" in nombre:
        return HOJA_ESCALERAS
    if "epp" in nombre or "proteccion personal" in nombre:
        return HOJA_EPP
    if "inalambric" in nombre or "bateria" in nombre:
        return HOJA_INALAMBRICAS
    if "cable" in nombre or ("electri" in nombre and "inalambric" not in nombre):
        return HOJA_CON_CABLE
    if "manual" in nombre:
        return HOJA_MANUALES
    return HOJA_MANUALES

def _fecha(valor):
    return valor.strftime("%d/%m/%Y") if valor else ""

def _codigo_documento(inspeccion_o_plantilla):
    """
    Devuelve el código fijo de documento SST del formato según la plantilla:
    - Manuales (plantilla "Manual"): FOR-SST-001
    - Eléctrica inalámbrica (plantilla "Inalámbrica"): FOR-SST-002
    - Eléctrica con cable (plantilla "Eléctrica"): FOR-SST-003
    - EPP (plantilla "EPP (equipo de protección personal)"): FOR-SST-004
    - Equipos de protección contra caídas: FOR-SST-005
    - Escaleras (plantilla "Escaleras"): FOR-SST-006
    """
    plantilla_nombre = ""
    if hasattr(inspeccion_o_plantilla, "plantilla") and inspeccion_o_plantilla.plantilla:
        plantilla_nombre = getattr(inspeccion_o_plantilla.plantilla, "nombre", "")
    elif hasattr(inspeccion_o_plantilla, "nombre"):
        plantilla_nombre = inspeccion_o_plantilla.nombre
    elif isinstance(inspeccion_o_plantilla, str):
        plantilla_nombre = inspeccion_o_plantilla

    nombre = _normalizar(plantilla_nombre)
    if "caida" in nombre or "caída" in nombre or "arnes" in nombre:
        return "FOR-SST-005"
    if "escalera" in nombre:
        return "FOR-SST-006"
    if "epp" in nombre or "proteccion personal" in nombre:
        return "FOR-SST-004"
    if "inalambric" in nombre or "bateria" in nombre:
        return "FOR-SST-002"
    if "cable" in nombre or ("electri" in nombre and "inalambric" not in nombre):
        return "FOR-SST-003"
    if "manual" in nombre:
        return "FOR-SST-001"

    insp_id = getattr(inspeccion_o_plantilla, "id", None)
    if insp_id:
        return f"FOR-SST-{insp_id:05d}"
    return "FOR-SST-001"

def _fecha_emision_hoy():
    from datetime import date
    return date.today().strftime("%d/%m/%Y")

# ─── EXCEL ─────────────────────────────────────────────────────────────────

def generar_excel_inspeccion(inspeccion):
    """
    Genera el formato Excel institucional completo con metadatos SST,
    color trimestral 5S, datos generales y firmas de conformidad.
    """
    return _generar_excel_simple(inspeccion)


    # Código del documento SST (H1) y fecha de emisión (H3)
    try:
        ws["H1"] = _codigo_documento(inspeccion)
        ws["H3"] = _fecha_emision_hoy()
    except Exception:
        pass

    if config["tipo"] == "grupal":
        ws[campos["tipo_herramienta"]] = inspeccion.material.nombre
        ws[campos["responsable"]] = inspector_nombre
        ws[campos["cant_inspeccionada"]] = inspeccion.cantidad_inspeccionada or ""
        ws[campos["fecha_inspeccion"]] = _fecha(inspeccion.fecha)
        ws[campos["cant_apta"]] = inspeccion.cantidad_apta or ""
        ws[campos["proxima_inspeccion"]] = _fecha(inspeccion.proxima_inspeccion)
        ws[campos["cant_no_apta"]] = inspeccion.cantidad_no_apta or ""
    else:
        ws[campos["codigo_herramienta"]] = codigo_objetivo
        ws[campos["proxima_inspeccion"]] = _fecha(inspeccion.proxima_inspeccion)
        ws[campos["marca"]] = inspeccion.material.marca
        ws[campos["inspector"]] = inspector_nombre
        ws[campos["fecha_inspeccion"]] = _fecha(inspeccion.fecha)
        ws[campos["nombre_herramienta"]] = inspeccion.material.nombre

    # Criterios: se apoya en el orden del criterio para ubicar la fila correcta.
    # Para plantillas "prestadas" (EPP, escaleras...) que usan una hoja cuya plantilla
    # original tiene criterios distintos, primero borramos los textos existentes en el
    # rango de datos y después escribimos los de la inspección.
    col_valor = {"cumple": "C", "no_cumple": "D", "no_aplica": "E"}
    fila_base = config["criterio_data_start"] - 1  # criterio orden=1 → primera fila de datos

    respuestas = list(inspeccion.respuestas.select_related("criterio").order_by("criterio__orden"))
    num_criterios_plantilla_nativa = config.get("num_criterios_nativos", 0)
    # Si la plantilla tiene más criterios fijos que las respuestas de esta inspección,
    # limpiamos los excedentes del template para que no queden textos fantasma.
    max_fila_usada = fila_base + len(respuestas)
    if num_criterios_plantilla_nativa > len(respuestas):
        for orden_extra in range(len(respuestas) + 1, num_criterios_plantilla_nativa + 1):
            fila_extra = fila_base + orden_extra
            for col_limpiar in ["A", "B", "C", "D", "E", "F"]:
                ws[f"{col_limpiar}{fila_extra}"] = None

    for resp in respuestas:
        fila = fila_base + resp.criterio.orden
        # Siempre sobreescribir número y texto del criterio (importante para plantillas no nativas)
        ws[f"A{fila}"] = resp.criterio.orden
        ws[f"B{fila}"] = resp.criterio.texto
        # Limpiar las tres columnas de valor antes de marcar
        for c in ["C", "D", "E"]:
            ws[f"{c}{fila}"] = None
        col = col_valor.get(resp.valor)
        if col:
            ws[f"{col}{fila}"] = "X"
        if resp.observacion:
            ws[f"F{fila}"] = resp.observacion

    if config["tipo"] == "grupal":
        aptas_row = config["resultado_aptas_row"]
        obs_row = config["resultado_observaciones_row"]
        if inspeccion.cantidad_no_apta:
            ws[f"A{aptas_row}"] = "\u2610 Todas las herramientas inspeccionadas se encuentran aptas."
            ws[f"A{obs_row}"] = "\u2611 Existen herramientas con observaciones (ver tabla anterior)."
        else:
            ws[f"A{aptas_row}"] = "\u2611 Todas las herramientas inspeccionadas se encuentran aptas."
            ws[f"A{obs_row}"] = "\u2610 Existen herramientas con observaciones (ver tabla anterior)."
    else:
        col_resultado = RESULTADO_COLS.get(inspeccion.resultado_general)
        if col_resultado:
            celda = ws[f"{col_resultado}{config['resultado_row']}"]
            celda.value = "X"
            celda.font = Font(bold=True, size=14)

        col_accion = ACCION_COLS.get(inspeccion.accion_tomada)
        if col_accion:
            celda = ws[f"{col_accion}{config['accion_row']}"]
            celda.value = "X"
            celda.font = Font(bold=True, size=14)

    if inspeccion.observaciones:
        ws[config["observaciones_generales"]] = inspeccion.observaciones

    # Insertar 5 filas de espacio en blanco antes de los cargos para que haya amplio espacio de firma manuscrita
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(row=r, column=1).value or "").upper()
        if "FIRMAS DE CONFORMIDAD" in val:
            ws.insert_rows(r + 1, 5)
            for empty_r in range(r + 1, r + 6):
                ws.row_dimensions[empty_r].height = 20
            break

    # Aplicar diseño oscuro institucional, bordes finos e insertar logo oficial
    _aplicar_estilo_oscuro_plantilla(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ─── Estilos y funciones auxiliares para Excel de Inspecciones ───────────────

EXCEL_HEADER_BG = "000000"         # Negro puro corporativo (combina 100% con el logo)
EXCEL_HEADER_TXT = "FFFFFF"
EXCEL_SUBHEADER_BG = "18181B"      # Negro grafito oscuro institucional
EXCEL_ROW_ALT = "F9FAFB"        # Fondo gris sutil para filas alternadas
EXCEL_ROW_BASE = "FFFFFF"       # Fondo base blanco
EXCEL_BORDER = "E5E7EB"         # Borde gris sutil
EXCEL_TEXT = "000000"           # Texto principal negro nítido

# Colores semánticos de resultado (versión saturada profesional con texto blanco)
EXCEL_APTA_BG = "065F46"        # Verde esmeralda oscuro
EXCEL_APTA_TXT = "FFFFFF"
EXCEL_REPARACION_BG = "92400E"  # Ámbar oscuro
EXCEL_REPARACION_TXT = "FFFFFF"
EXCEL_FUERA_BG = "991B1B"       # Rojo borgoña oscuro
EXCEL_FUERA_TXT = "FFFFFF"


def _get_logo_path():
    """Retorna la ruta al logo institucional PNG disponible."""
    from pathlib import Path
    p0 = Path(__file__).resolve().parent / "logo_incalpaca_header.png"
    p1 = Path(__file__).resolve().parent / "logo_incalpaca.png"
    p2 = Path(__file__).resolve().parent.parent.parent.parent / "frontend" / "public" / "logo-incalpaca.png"
    p3 = Path(__file__).resolve().parent.parent / "workorders" / "logo_brand.png"
    for p in [p0, p1, p2, p3]:
        if p.exists():
            return p
    return None


def _insert_logo(ws, cell="A1", width=125, height=37):
    """Inserta la imagen del logo institucional en la celda indicada."""
    from openpyxl.drawing.image import Image as OpenpyxlImage
    p = _get_logo_path()
    if p:
        try:
            img = OpenpyxlImage(str(p))
            img.width = width
            img.height = height
            ws.add_image(img, cell)
        except Exception:
            pass


def _aplicar_estilo_oscuro_plantilla(ws):
    """
    Aplica diseño corporativo en negro estricto e inserta el logo oficial en la plantilla.
    Reemplaza todos los fondos celestes pastel por la paleta institucional (#000000 / #18181B).
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # 1. Cabecera ejecutiva integrada en negro puro (Filas 1 a 5)
    dark_fill = _excel_fill("000000")
    slate_fill = _excel_fill("18181B")
    header_border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )

    for r in range(1, 6):
        # A..E: Bloque del Logo y Título
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = dark_fill
            cell.border = header_border
            if cell.value:
                cell.font = Font(bold=True, size=11 if r <= 2 else 9.5, color="FFFFFF", name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # F..H: Caja de Metadatos
        for c in range(6, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = header_border
            if c in [6, 7]:
                cell.fill = slate_fill
                cell.font = Font(bold=True, size=8.5, color="A1A1AA", name="Calibri")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.fill = dark_fill
                cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Insertar Logo oficial horizontal en A1
    _insert_logo(ws, cell="A1", width=125, height=37)

    # 2. Paleta institucional para el resto de la hoja
    fill_section = _excel_fill("000000")      # Negro puro
    fill_subhead = _excel_fill("18181B")      # Negro zinc grafito
    font_section = Font(bold=True, size=10.5, color="FFFFFF", name="Calibri")
    font_subhead = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
    thin_border = _excel_thin_border()

    for r in range(7, ws.max_row + 1):
        c1_val = str(ws.cell(row=r, column=1).value or "").strip().upper()

        # A) Títulos de sección principales
        if any(keyword in c1_val for keyword in [
            "DATOS GENERALES",
            "CRITERIOS DE INSPECCIÓN",
            "CRITERIOS DE INSPECCION",
            "HERRAMIENTAS CON OBSERVACIONES",
            "RESULTADO FINAL",
            "OBSERVACIONES GENERALES",
            "FIRMAS DE CONFORMIDAD",
        ]):
            ws.row_dimensions[r].height = 25
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_section
                cell.font = font_section
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            continue

        # B) Encabezados de columnas de tablas (N°, Criterio de inspección, Cumple, Código...)
        if any(keyword in c1_val for keyword in ["N°", "Nº", "N.", "CÓDIGO", "CODIGO"]):
            ws.row_dimensions[r].height = 22
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_subhead
                cell.font = font_subhead
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            continue

        # C) Bloque de datos generales (R8..R12)
        if 8 <= r <= 12 and "CRITERIO" not in c1_val:
            ws.row_dimensions[r].height = 20
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None or ws.cell(row=r, column=1).value:
                    cell.border = thin_border
                    if c in [1, 5] and str(cell.value or "").strip().endswith(":"):
                        cell.font = Font(bold=True, size=9.5, color="1E293B", name="Calibri")
                        cell.fill = _excel_fill("F8FAFC")
            continue

        # D) Filas de criterios de evaluación: centrar marcas 'X', alternar zebra striping
        if str(ws.cell(row=r, column=1).value or "").strip().isdigit():
            ws.row_dimensions[r].height = 20
            idx_criterio = int(str(ws.cell(row=r, column=1).value).strip())
            is_alt = (idx_criterio % 2 == 0)
            row_bg = _excel_fill(EXCEL_ROW_ALT if is_alt else EXCEL_ROW_BASE)
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = row_bg
                cell.border = thin_border
                if c in [1, 3, 4, 5]:  # N°, Cumple, No cumple, No aplica
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if str(cell.value or "").strip().upper() == "X":
                        cell.font = Font(bold=True, size=11, color="0F172A", name="Calibri")
                elif c == 2:  # Texto criterio
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.font = Font(size=9.5, color="0F172A", name="Calibri")

# Colores semánticos de resultado (versión saturada profesional con texto blanco)
EXCEL_APTA_BG = "065F46"        # Verde esmeralda oscuro
EXCEL_APTA_TXT = "FFFFFF"
EXCEL_REPARACION_BG = "92400E"  # Ámbar oscuro
EXCEL_REPARACION_TXT = "FFFFFF"
EXCEL_FUERA_BG = "991B1B"       # Rojo borgoña oscuro
EXCEL_FUERA_TXT = "FFFFFF"


def _excel_fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _excel_thin_border():
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color=EXCEL_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _excel_header_font(bold=True, color=EXCEL_HEADER_TXT, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _excel_normal_font(bold=False, size=10, color=EXCEL_TEXT):
    return Font(bold=bold, size=size, color=color, name="Calibri")


def _excel_write_header_row(ws, row, columns, bg_color=EXCEL_HEADER_BG, font_color=EXCEL_HEADER_TXT):
    from openpyxl.styles import Alignment
    fill = _excel_fill(bg_color)
    font = _excel_header_font(bold=True, color=font_color, size=11)
    ws.row_dimensions[row].height = 26
    for col_idx, text in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _excel_thin_border()


def _excel_write_data_row(ws, row, values, row_fill=None, is_alt=False, alignments=None, cell_styles=None):
    from openpyxl.styles import Alignment
    bg = row_fill or (EXCEL_ROW_ALT if is_alt else EXCEL_ROW_BASE)
    fill = _excel_fill(bg) if bg else None
    ws.row_dimensions[row].height = 20
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = _excel_thin_border()
        if cell_styles and col_idx in cell_styles:
            c_bg, c_fc, c_bold = cell_styles[col_idx]
            cell.fill = _excel_fill(c_bg)
            cell.font = _excel_normal_font(bold=c_bold, color=c_fc, size=10)
        else:
            if fill:
                cell.fill = fill
            cell.font = _excel_normal_font(bold=False, color=EXCEL_TEXT, size=10)

        align = alignments.get(col_idx, "left") if alignments else "left"
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)


def _excel_set_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _excel_freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _generar_excel_simple(inspeccion):
    """
    Genera el formato Excel institucional completo de Check List de Inspección
    con encabezado formal SST, metadatos, color trimestral 5S, datos generales
    completos, evaluación de criterios y 4 firmas de conformidad.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    from apps.inspeccion.utils import color_inspeccion_actual

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspección"

    objetivo = (
        inspeccion.pieza.codigo
        if getattr(inspeccion, "pieza", None)
        else (inspeccion.material.codigo if getattr(inspeccion, "material", None) else "—")
    )
    mat_nombre = inspeccion.material.nombre if getattr(inspeccion, "material", None) else "—"
    codigo_doc = _codigo_documento(inspeccion)
    fecha_emision = _fecha_emision_hoy()
    inspector_nombre = (
        inspeccion.inspector.get_full_name() or inspeccion.inspector.username
        if getattr(inspeccion, "inspector", None)
        else "N/A"
    )
    modalidad_label = "Planificada" if getattr(inspeccion, "modalidad", "planificada") == "planificada" else "No planificada"
    frecuencia_label = (getattr(inspeccion, "frecuencia", "trimestral") or "trimestral").capitalize()
    area_label = getattr(inspeccion, "area_trabajo", "") or "Facility Management"
    referencia_ot = getattr(inspeccion, "referencia_orden", "") or "—"
    color_info = color_inspeccion_actual(para_fecha=inspeccion.fecha.date() if getattr(inspeccion, "fecha", None) else None)
    color_actual = color_info["actual"]

    # 1. Encabezado principal: Logo Incalpaca FM a la izquierda (A1:B5) | Título central (C1:E5) | Metadatos SST (F1:H5)
    ws.merge_cells("A1:B5")
    logo_cell = ws["A1"]
    logo_cell.fill = _excel_fill("FFFFFF")
    logo_cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(1, 6):
        for col_t in ["A", "B"]:
            ws[f"{col_t}{r}"].fill = _excel_fill("FFFFFF")
            ws[f"{col_t}{r}"].border = _excel_thin_border()

    try:
        from pathlib import Path
        from openpyxl.drawing.image import Image as OpenpyxlImage
        header_logo_path = Path(__file__).resolve().parent / "logo_incalpaca_header.png"
        if not header_logo_path.exists():
            header_logo_path = Path(__file__).resolve().parent / "logo_incalpaca.png"
        if header_logo_path.exists():
            hdr_img = OpenpyxlImage(str(header_logo_path))
            hdr_img.width = 175
            hdr_img.height = 35
            ws.add_image(hdr_img, "A2")
        else:
            logo_cell.value = "INCALPACA | FM"
            logo_cell.font = Font(bold=True, size=11, color="000000", name="Calibri")
    except Exception:
        logo_cell.value = "INCALPACA | FM"
        logo_cell.font = Font(bold=True, size=11, color="000000", name="Calibri")

    hoja_tipo = _detectar_hoja(inspeccion.plantilla.nombre if getattr(inspeccion, "plantilla", None) else "")

    if hoja_tipo == HOJA_CAIDAS:
        nombre_formato = "FORMATO DE INSPECCIÓN DE EQUIPOS DE PROTECCIÓN CONTRA CAÍDAS"
        proceso_label = "Inspección de Equipos Contra Caídas"
        criterios_header_label = "CRITERIOS DE INSPECCIÓN DE EQUIPOS DE PROTECCIÓN CONTRA CAÍDAS"
    elif hoja_tipo == HOJA_EPP:
        nombre_formato = "FORMATO DE INSPECCIÓN DE EQUIPO DE PROTECCIÓN PERSONAL (EPP)"
        proceso_label = "Inspección de Equipos de Protección Personal"
        criterios_header_label = "CRITERIOS DE INSPECCIÓN DE EPP (CASCO, ANTEOJOS, GUANTES, ZAPATOS, CARETAS, PROTECTORES AUDITIVOS)"
    elif hoja_tipo == HOJA_ESCALERAS:
        nombre_formato = "FORMATO DE INSPECCIÓN DE ESCALERAS Y EQUIPOS DE ALTURA"
        proceso_label = "Inspección de Escaleras y Equipos de Altura"
        criterios_header_label = "CRITERIOS DE INSPECCIÓN DE ESCALERAS"
    elif hoja_tipo == HOJA_INALAMBRICAS:
        nombre_formato = "FORMATO DE INSPECCIÓN DE HERRAMIENTAS ELÉCTRICAS INALÁMBRICAS"
        proceso_label = "Inspección de Herramientas Eléctricas Inalámbricas"
        criterios_header_label = "CRITERIOS DE INSPECCIÓN VISUAL, ELÉCTRICA Y FUNCIONAL (HERRAMIENTA Y BATERÍA)"
    elif hoja_tipo == HOJA_CON_CABLE:
        nombre_formato = "FORMATO DE INSPECCIÓN DE HERRAMIENTAS ELÉCTRICAS CON CABLE"
        proceso_label = "Inspección de Herramientas Eléctricas con Cable"
        criterios_header_label = "CRITERIOS DE INSPECCIÓN VISUAL, ELÉCTRICA Y FUNCIONAL"
    else:
        nombre_formato = "FORMATO DE INSPECCIÓN GRUPAL DE HERRAMIENTAS MANUALES" if getattr(inspeccion, "tipo", "individual") == "grupal" else "FORMATO DE INSPECCIÓN DE HERRAMIENTAS MANUALES"
        proceso_label = "Inspección de Herramientas Manuales"
        criterios_header_label = "CRITERIOS DE INSPECCIÓN (APLICABLES AL GRUPO DE HERRAMIENTAS INSPECCIONADO)" if getattr(inspeccion, "tipo", "individual") == "grupal" else "CRITERIOS DE INSPECCIÓN DE HERRAMIENTAS MANUALES"

    # Título central C1:E5 (Título del formato + Área)
    ws.merge_cells("C1:E5")
    title_cell = ws["C1"]
    title_cell.value = f"{nombre_formato}\nÁrea: {area_label}"
    title_cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    title_cell.fill = _excel_fill("000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(1, 6):
        for col_t in ["C", "D", "E"]:
            ws[f"{col_t}{r}"].fill = _excel_fill("000000")
            ws[f"{col_t}{r}"].border = _excel_thin_border()

    for r in range(1, 6):
        ws.row_dimensions[r].height = 18

    # 2. Bloque de Metadatos SST (F1:H5)
    meta_cabecera = [
        ("Código:", codigo_doc),
        ("Versión:", "01"),
        ("Fecha de emisión:", fecha_emision),
        ("Frecuencia:", frecuencia_label),
        ("Proceso:", proceso_label),
    ]
    for r_offset, (lbl, val) in enumerate(meta_cabecera, start=1):
        ws.merge_cells(start_row=r_offset, start_column=6, end_row=r_offset, end_column=7)
        c_lbl = ws.cell(row=r_offset, column=6, value=lbl)
        c_lbl.font = Font(bold=True, size=8.5, color="A1A1AA", name="Calibri")
        c_lbl.fill = _excel_fill("18181B")
        c_lbl.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r_offset, column=7).fill = _excel_fill("18181B")

        c_val = ws.cell(row=r_offset, column=8, value=val)
        c_val.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        c_val.fill = _excel_fill("000000")
        c_val.alignment = Alignment(horizontal="left", vertical="center")

        for col_i in range(6, 9):
            ws.cell(row=r_offset, column=col_i).border = _excel_thin_border()

    # 3. Sección DATOS GENERALES DE LA INSPECCIÓN (Fila 7)
    ws.merge_cells("A7:H7")
    sec1 = ws["A7"]
    sec1.value = "DATOS GENERALES DE LA INSPECCIÓN"
    sec1.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    sec1.fill = _excel_fill(EXCEL_HEADER_BG)
    sec1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 22
    for c in range(1, 9):
        ws.cell(row=7, column=c).border = _excel_thin_border()

    # Cantidades individuales vs grupales
    if getattr(inspeccion, "tipo", "individual") == "grupal":
        cant_insp_str = str(inspeccion.cantidad_inspeccionada if inspeccion.cantidad_inspeccionada is not None else 1)
        cant_apta_num = inspeccion.cantidad_apta if inspeccion.cantidad_apta is not None else 0
        cant_no_apta_num = inspeccion.cantidad_no_apta if inspeccion.cantidad_no_apta is not None else 0
        cant_resumen = f"{cant_apta_num} aptas / {cant_no_apta_num} no aptas"
    else:
        cant_insp_str = "1 (Individual)"
        if inspeccion.cantidad_apta is not None and inspeccion.cantidad_no_apta is not None:
            cant_apta_num = inspeccion.cantidad_apta
            cant_no_apta_num = inspeccion.cantidad_no_apta
        else:
            cant_apta_num = 1 if getattr(inspeccion, "resultado_general", "apta") == "apta" else 0
            cant_no_apta_num = 0 if getattr(inspeccion, "resultado_general", "apta") == "apta" else 1
        cant_resumen = f"{cant_apta_num} apta{'s' if cant_apta_num != 1 else ''} / {cant_no_apta_num} no apta{'s' if cant_no_apta_num != 1 else ''}"

    datos_generales_rows = [
        [("Tipo de herramienta / Material:", True), (mat_nombre, False), ("Responsable:", True), (inspector_nombre, False)],
        [("Tipo de inspección:", True), (modalidad_label, False), ("Fecha de inspección:", True), (_fecha(inspeccion.fecha), False)],
        [("Frecuencia planificada:", True), (frecuencia_label, False), ("Próxima inspección:", True), (_fecha(inspeccion.proxima_inspeccion), False)],
        [("Área de trabajo / Lugar:", True), (area_label, False), ("Referencia (OT/OL/OP):", True), (referencia_ot, False)],
        [("Cantidad inspeccionada:", True), (cant_insp_str, False), ("Cantidad apta / no apta:", True), (cant_resumen, False)],
    ]
    for r_offset, (lbl1, val1, lbl2, val2) in enumerate(datos_generales_rows, start=8):
        ws.row_dimensions[r_offset].height = 20
        # Columna 1 (A:B)
        ws.merge_cells(start_row=r_offset, start_column=1, end_row=r_offset, end_column=2)
        c_lbl1 = ws.cell(row=r_offset, column=1, value=lbl1[0])
        c_lbl1.font = Font(bold=True, size=9, color="334155", name="Calibri")
        c_lbl1.fill = _excel_fill("F8FAFC")
        for col_i in range(1, 3):
            ws.cell(row=r_offset, column=col_i).border = _excel_thin_border()

        # Columna 2 (C:D)
        ws.merge_cells(start_row=r_offset, start_column=3, end_row=r_offset, end_column=4)
        c_val1 = ws.cell(row=r_offset, column=3, value=val1[0])
        c_val1.font = Font(bold=False, size=9.5, color="0F172A", name="Calibri")
        c_val1.fill = _excel_fill("FFFFFF")
        for col_i in range(3, 5):
            ws.cell(row=r_offset, column=col_i).border = _excel_thin_border()

        # Columna 3 (E:F)
        ws.merge_cells(start_row=r_offset, start_column=5, end_row=r_offset, end_column=6)
        c_lbl2 = ws.cell(row=r_offset, column=5, value=lbl2[0])
        c_lbl2.font = Font(bold=True, size=9, color="334155", name="Calibri")
        c_lbl2.fill = _excel_fill("F8FAFC")
        for col_i in range(5, 7):
            ws.cell(row=r_offset, column=col_i).border = _excel_thin_border()

        # Columna 4 (G:H)
        ws.merge_cells(start_row=r_offset, start_column=7, end_row=r_offset, end_column=8)
        c_val2 = ws.cell(row=r_offset, column=7, value=val2[0])
        c_val2.font = Font(bold=False, size=9.5, color="0F172A", name="Calibri")
        c_val2.fill = _excel_fill("FFFFFF")
        for col_i in range(7, 9):
            ws.cell(row=r_offset, column=col_i).border = _excel_thin_border()

    # 4. Sección CÓDIGO TRIMESTRAL / BIMESTRAL DE INSPECCIÓN (COLOR DEL MES - SISTEMA 5S)
    frecuencia_val = (getattr(inspeccion, "frecuencia", "") or "").lower()
    es_bimestral = (
        frecuencia_val == "bimestral"
        or hoja_tipo in [HOJA_INALAMBRICAS, HOJA_CON_CABLE]
    )

    titulo_color_sec = "CÓDIGO BIMESTRAL DE INSPECCIÓN (COLOR DEL MES - SISTEMA 5S)" if es_bimestral else "CÓDIGO TRIMESTRAL DE INSPECCIÓN (COLOR DEL MES - SISTEMA 5S)"

    ws.merge_cells("A13:H13")
    sec_color = ws["A13"]
    sec_color.value = titulo_color_sec
    sec_color.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
    sec_color.fill = _excel_fill(EXCEL_SUBHEADER_BG)
    sec_color.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[13].height = 20
    for c in range(1, 9):
        ws.cell(row=13, column=c).border = _excel_thin_border()

    fecha_base_ex = inspeccion.fecha.date() if getattr(inspeccion, "fecha", None) else date.today()
    mes_actual_ex = fecha_base_ex.month

    if es_bimestral:
        periodo_activo_num = ((mes_actual_ex - 1) // 2) + 1  # 1 a 6
        filas_periodos = [
            {"num": 1, "label": "I Bimestre",   "meses": "Enero - Febrero",         "color_nombre": "Amarillo", "hex": "FFFF00", "txt_color": "000000"},
            {"num": 2, "label": "II Bimestre",  "meses": "Marzo - Abril",           "color_nombre": "Verde",    "hex": "00B050", "txt_color": "FFFFFF"},
            {"num": 3, "label": "III Bimestre", "meses": "Mayo - Junio",            "color_nombre": "Rojo",     "hex": "FF0000", "txt_color": "FFFFFF"},
            {"num": 4, "label": "IV Bimestre",  "meses": "Julio - Agosto",          "color_nombre": "Azul",     "hex": "0070C0", "txt_color": "FFFFFF"},
            {"num": 5, "label": "V Bimestre",   "meses": "Septiembre - Octubre",    "color_nombre": "Negro",    "hex": "000000", "txt_color": "FFFFFF"},
            {"num": 6, "label": "VI Bimestre",  "meses": "Noviembre - Diciembre",   "color_nombre": "Blanco",   "hex": "FFFFFF", "txt_color": "000000"},
        ]
    else:
        periodo_activo_num = ((mes_actual_ex - 1) // 3) + 1  # 1 a 4
        filas_periodos = [
            {"num": 1, "label": "I Trimestre",   "meses": "Enero - Febrero - Marzo",         "color_nombre": "Amarillo", "hex": "FFFF00", "txt_color": "000000"},
            {"num": 2, "label": "II Trimestre",  "meses": "Abril - Mayo - Junio",            "color_nombre": "Verde",    "hex": "00B050", "txt_color": "FFFFFF"},
            {"num": 3, "label": "III Trimestre", "meses": "Julio - Agosto - Septiembre",     "color_nombre": "Azul",     "hex": "0070C0", "txt_color": "FFFFFF"},
            {"num": 4, "label": "IV Trimestre",  "meses": "Octubre - Noviembre - Diciembre", "color_nombre": "Rojo",     "hex": "FF0000", "txt_color": "FFFFFF"},
        ]

    for idx_t, per in enumerate(filas_periodos, start=14):
        ws.row_dimensions[idx_t].height = 16.5
        es_activo = (per["num"] == periodo_activo_num)

        border_style = "medium" if es_activo else "thin"
        border_color = "000000" if es_activo else "CBD5E1"
        row_border = Border(
            left=Side(style=border_style, color=border_color),
            right=Side(style=border_style, color=border_color),
            top=Side(style=border_style, color=border_color),
            bottom=Side(style=border_style, color=border_color),
        )
        bg_fila = "FEF08A" if (es_activo and per["hex"] == "FFFF00") else ("F1F5F9" if es_activo else "FFFFFF")

        # Periodo (Cols A:B)
        ws.merge_cells(start_row=idx_t, start_column=1, end_row=idx_t, end_column=2)
        txt_per = f"{'▶ ' if es_activo else '   '}{per['label']}{' (VIGENTE)' if es_activo else ''}"
        c_per = ws.cell(row=idx_t, column=1, value=txt_per)
        c_per.font = Font(bold=es_activo, size=8.5, color="0F172A", name="Calibri")
        c_per.fill = _excel_fill(bg_fila)
        c_per.alignment = Alignment(horizontal="left", vertical="center")
        for col_k in [1, 2]:
            ws.cell(row=idx_t, column=col_k).border = row_border
            ws.cell(row=idx_t, column=col_k).fill = _excel_fill(bg_fila)

        # Meses (Cols C:F)
        ws.merge_cells(start_row=idx_t, start_column=3, end_row=idx_t, end_column=6)
        c_mes = ws.cell(row=idx_t, column=3, value=per["meses"])
        c_mes.font = Font(bold=es_activo, size=8.5, color="0F172A", name="Calibri")
        c_mes.fill = _excel_fill(bg_fila)
        c_mes.alignment = Alignment(horizontal="left", vertical="center")
        for col_k in range(3, 7):
            ws.cell(row=idx_t, column=col_k).border = row_border
            ws.cell(row=idx_t, column=col_k).fill = _excel_fill(bg_fila)

        # Color (Cols G:H)
        ws.merge_cells(start_row=idx_t, start_column=7, end_row=idx_t, end_column=8)
        c_col = ws.cell(row=idx_t, column=7, value=per["color_nombre"].upper())
        c_col.font = Font(bold=True, size=8.5, color=per["txt_color"], name="Calibri")
        c_col.fill = _excel_fill(per["hex"])
        c_col.alignment = Alignment(horizontal="center", vertical="center")
        for col_k in [7, 8]:
            ws.cell(row=idx_t, column=col_k).border = row_border
            ws.cell(row=idx_t, column=col_k).fill = _excel_fill(per["hex"])

    next_r = 14 + len(filas_periodos)
    ws.row_dimensions[next_r].height = 4  # espaciador
    next_r += 1

    # 5. Sección TIPO DE HERRAMIENTAS MANUALES — SOLO si es manual y se seleccionaron tipos
    tipos_marcados = getattr(inspeccion, "tipos_herramientas", []) or []
    if hoja_tipo == HOJA_MANUALES and len(tipos_marcados) > 0:
        ws.merge_cells(start_row=next_r, start_column=1, end_row=next_r, end_column=8)
        sec_tipos = ws.cell(row=next_r, column=1, value="TIPO DE HERRAMIENTAS MANUALES (Marcar las que aplican a la inspección)")
        sec_tipos.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
        sec_tipos.fill = _excel_fill(EXCEL_SUBHEADER_BG)
        sec_tipos.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[next_r].height = 20
        for c in range(1, 9):
            ws.cell(row=next_r, column=c).border = _excel_thin_border()

        grupos_h = [
            ("Herramientas de golpe", "Herramientas de medición"),
            ("Herramientas de corte", "Herramientas de sujeción"),
            ("Herramientas de cohesión", "Herramientas de pintura"),
            ("Herramientas de torsión y ajuste", "Otras herramientas"),
        ]
        for offset_g, (g1, g2) in enumerate(grupos_h, start=1):
            r_g = next_r + offset_g
            ws.row_dimensions[r_g].height = 19
            marca1 = "[X]" if g1 in tipos_marcados else "[  ]"
            ws.merge_cells(start_row=r_g, start_column=1, end_row=r_g, end_column=4)
            c_g1 = ws.cell(row=r_g, column=1, value=f"  {marca1}  {g1}")
            c_g1.font = Font(bold=marca1 == "[X]", size=9, color="0F172A", name="Calibri")
            c_g1.fill = _excel_fill("FFFFFF" if marca1 != "[X]" else "FEF3C7")
            c_g1.alignment = Alignment(horizontal="left", vertical="center")
            for col_i in range(1, 5):
                ws.cell(row=r_g, column=col_i).border = _excel_thin_border()

            marca2 = "[X]" if g2 in tipos_marcados else "[  ]"
            ws.merge_cells(start_row=r_g, start_column=5, end_row=r_g, end_column=8)
            c_g2 = ws.cell(row=r_g, column=5, value=f"  {marca2}  {g2}")
            c_g2.font = Font(bold=marca2 == "[X]", size=9, color="0F172A", name="Calibri")
            c_g2.fill = _excel_fill("FFFFFF" if marca2 != "[X]" else "FEF3C7")
            c_g2.alignment = Alignment(horizontal="left", vertical="center")
            for col_i in range(5, 9):
                ws.cell(row=r_g, column=col_i).border = _excel_thin_border()

        next_r = next_r + 5
        ws.row_dimensions[next_r].height = 4  # espaciador
        header_row = next_r + 1
    else:
        header_row = next_r

    # 6. Tabla de Criterios
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=8)
    sec_crit_title = ws.cell(row=header_row, column=1, value=criterios_header_label)
    sec_crit_title.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
    sec_crit_title.fill = _excel_fill(EXCEL_HEADER_BG)
    sec_crit_title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[header_row].height = 20
    for c in range(1, 9):
        ws.cell(row=header_row, column=c).border = _excel_thin_border()

    header_row += 1
    ws.row_dimensions[header_row].height = 22
    ws.cell(row=header_row, column=1, value="N°")
    ws.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=4)

    if hoja_tipo == HOJA_CAIDAS:
        ws.cell(row=header_row, column=2, value="Criterio de inspección (Defectos a evaluar)")
        ws.cell(row=header_row, column=5, value="SÍ")
        ws.cell(row=header_row, column=6, value="NO")
    else:
        ws.cell(row=header_row, column=2, value="Criterio de inspección")
        ws.cell(row=header_row, column=5, value="Cumple")
        ws.cell(row=header_row, column=6, value="No cumple")

    ws.cell(row=header_row, column=7, value="No aplica")
    ws.cell(row=header_row, column=8, value="Observaciones")

    for c_i in range(1, 9):
        cell = ws.cell(row=header_row, column=c_i)
        cell.fill = _excel_fill(EXCEL_SUBHEADER_BG)
        cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        cell.border = _excel_thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    plantilla_obj = (
        inspeccion.plantilla
        if getattr(inspeccion, "plantilla", None)
        else (inspeccion.material.subcategoria.plantilla_inspeccion if getattr(inspeccion, "material", None) and getattr(inspeccion.material, "subcategoria", None) else None)
    )
    if not plantilla_obj:
        from apps.inspeccion.models import PlantillaCriterio
        nombre_buscar = "Inalámbrica" if hoja_tipo == HOJA_INALAMBRICAS else ("Eléctrica" if hoja_tipo == HOJA_CON_CABLE else ("Manual" if hoja_tipo == HOJA_MANUALES else ("EPP (equipo de protección personal)" if hoja_tipo == HOJA_EPP else ("Escaleras" if hoja_tipo == HOJA_ESCALERAS else "Equipos de protección contra caídas"))))
        plantilla_obj = PlantillaCriterio.objects.filter(nombre__icontains=nombre_buscar.split()[0]).first()

    criterios_oficiales = list(plantilla_obj.criterios.all().order_by("orden")) if plantilla_obj else []
    respuestas_dict = {r.criterio_id: r for r in inspeccion.respuestas.select_related("criterio").all()}

    filas_criterios_render = []
    if criterios_oficiales:
        for crit in criterios_oficiales:
            resp = respuestas_dict.get(crit.id)
            valor_resp = resp.valor if resp else ("cumple" if getattr(inspeccion, "resultado_general", "apta") == "apta" else "")
            obs_resp = resp.observacion if resp else ""
            filas_criterios_render.append((crit.orden, crit.texto, valor_resp, obs_resp))
    elif respuestas_dict:
        for r in sorted(respuestas_dict.values(), key=lambda x: x.criterio.orden):
            filas_criterios_render.append((r.criterio.orden, r.criterio.texto, r.valor, r.observacion or ""))

    for idx, (crit_orden, crit_texto, resp_valor, resp_obs) in enumerate(filas_criterios_render, start=1):
        r_num = header_row + idx
        ws.row_dimensions[r_num].height = 20
        is_alt = (idx % 2 == 0)
        bg = EXCEL_ROW_ALT if is_alt else EXCEL_ROW_BASE

        c_ord = ws.cell(row=r_num, column=1, value=crit_orden)
        c_ord.alignment = Alignment(horizontal="center", vertical="center")
        c_ord.font = Font(bold=True, size=9.5, color="0F172A", name="Calibri")

        ws.merge_cells(start_row=r_num, start_column=2, end_row=r_num, end_column=4)
        c_txt = ws.cell(row=r_num, column=2, value=crit_texto)
        c_txt.font = Font(size=9.5, color="0F172A", name="Calibri")
        c_txt.alignment = Alignment(horizontal="left", vertical="center")

        if hoja_tipo == HOJA_CAIDAS:
            # En caídas: SÍ = defecto presente (no_cumple), NO = sin defecto (cumple)
            ws.cell(row=r_num, column=5, value="X" if resp_valor == "no_cumple" else "").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r_num, column=6, value="X" if resp_valor == "cumple" else "").alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.cell(row=r_num, column=5, value="X" if resp_valor == "cumple" else "").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r_num, column=6, value="X" if resp_valor == "no_cumple" else "").alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=r_num, column=7, value="X" if resp_valor == "no_aplica" else "").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_num, column=8, value=resp_obs).alignment = Alignment(horizontal="left", vertical="center")

        for col_i in range(1, 9):
            cell = ws.cell(row=r_num, column=col_i)
            cell.border = _excel_thin_border()
            cell.fill = _excel_fill(bg)
            if col_i in [5, 6, 7] and cell.value == "X":
                cell.font = Font(bold=True, size=11, color="0F172A", name="Calibri")

    curr_row = header_row + len(filas_criterios_render)

    # 6.5. Tabla de Herramientas / EPP con Observaciones — SOLO para Manuales (FOR-SST-001) y EPP (FOR-SST-004)
    if hoja_tipo in (HOJA_MANUALES, HOJA_EPP):
        items_obs = list(inspeccion.items_con_observacion.all())
        curr_row += 1
        ws.row_dimensions[curr_row].height = 4  # espaciador
        curr_row += 1

        if hoja_tipo == HOJA_EPP:
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
            sec_obs_title = ws.cell(row=curr_row, column=1, value="EPP CON OBSERVACIONES (registrar únicamente las que presenten condición insegura)")
            sec_obs_title.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
            sec_obs_title.fill = _excel_fill(EXCEL_HEADER_BG)
            sec_obs_title.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[curr_row].height = 20
            for c in range(1, 9):
                ws.cell(row=curr_row, column=c).border = _excel_thin_border()

            curr_row += 1
            ws.row_dimensions[curr_row].height = 20
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
            ws.cell(row=curr_row, column=1, value="Código")
            ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
            ws.cell(row=curr_row, column=3, value="Nombre del EPP")
            ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=8)
            ws.cell(row=curr_row, column=5, value="Observación encontrada")

            for c_i in range(1, 9):
                cell = ws.cell(row=curr_row, column=c_i)
                cell.fill = _excel_fill(EXCEL_SUBHEADER_BG)
                cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
                cell.border = _excel_thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")

            filas_obs = items_obs if len(items_obs) > 0 else [None, None]
            for obs_item in filas_obs:
                curr_row += 1
                ws.row_dimensions[curr_row].height = 19
                cod_val = obs_item.codigo if obs_item else ""
                nom_val = obs_item.nombre if obs_item else ""
                obs_val = obs_item.observacion_encontrada if obs_item else ""

                ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
                c_cod = ws.cell(row=curr_row, column=1, value=cod_val)
                c_cod.alignment = Alignment(horizontal="center", vertical="center")
                c_cod.font = Font(size=9, color="0F172A", name="Calibri")

                ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
                c_nom = ws.cell(row=curr_row, column=3, value=nom_val)
                c_nom.alignment = Alignment(horizontal="left", vertical="center")
                c_nom.font = Font(size=9, color="0F172A", name="Calibri")

                ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=8)
                c_obs = ws.cell(row=curr_row, column=5, value=obs_val)
                c_obs.alignment = Alignment(horizontal="left", vertical="center")
                c_obs.font = Font(size=9, color="0F172A", name="Calibri")

                for col_i in range(1, 9):
                    ws.cell(row=curr_row, column=col_i).border = _excel_thin_border()
                    ws.cell(row=curr_row, column=col_i).fill = _excel_fill("FFFFFF")

        else:
            # HOJA_MANUALES
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
            sec_obs_title = ws.cell(row=curr_row, column=1, value="HERRAMIENTAS CON OBSERVACIONES (registrar únicamente las que presenten condición insegura)")
            sec_obs_title.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
            sec_obs_title.fill = _excel_fill(EXCEL_HEADER_BG)
            sec_obs_title.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[curr_row].height = 20
            for c in range(1, 9):
                ws.cell(row=curr_row, column=c).border = _excel_thin_border()

            curr_row += 1
            ws.row_dimensions[curr_row].height = 20
            ws.cell(row=curr_row, column=1, value="Código")
            ws.cell(row=curr_row, column=2, value="Nombre de la herramienta")
            ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
            ws.cell(row=curr_row, column=3, value="Observación encontrada")
            ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
            ws.cell(row=curr_row, column=5, value="Acción recomendada")
            ws.merge_cells(start_row=curr_row, start_column=7, end_row=curr_row, end_column=8)
            ws.cell(row=curr_row, column=7, value="Estado")

            for c_i in range(1, 9):
                cell = ws.cell(row=curr_row, column=c_i)
                cell.fill = _excel_fill(EXCEL_SUBHEADER_BG)
                cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
                cell.border = _excel_thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")

            filas_obs = items_obs if len(items_obs) > 0 else [None, None]
            for obs_item in filas_obs:
                curr_row += 1
                ws.row_dimensions[curr_row].height = 19
                cod_val = obs_item.codigo if obs_item else ""
                nom_val = obs_item.nombre if obs_item else ""
                obs_val = obs_item.observacion_encontrada if obs_item else ""
                acc_val = obs_item.accion_recomendada if obs_item else ""
                est_val = obs_item.estado if obs_item else ""

                c_cod = ws.cell(row=curr_row, column=1, value=cod_val)
                c_cod.alignment = Alignment(horizontal="center", vertical="center")
                c_cod.font = Font(size=9, color="0F172A", name="Calibri")

                c_nom = ws.cell(row=curr_row, column=2, value=nom_val)
                c_nom.alignment = Alignment(horizontal="left", vertical="center")
                c_nom.font = Font(size=9, color="0F172A", name="Calibri")

                ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
                c_obs = ws.cell(row=curr_row, column=3, value=obs_val)
                c_obs.alignment = Alignment(horizontal="left", vertical="center")
                c_obs.font = Font(size=9, color="0F172A", name="Calibri")

                ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
                c_acc = ws.cell(row=curr_row, column=5, value=acc_val)
                c_acc.alignment = Alignment(horizontal="left", vertical="center")
                c_acc.font = Font(size=9, color="0F172A", name="Calibri")

                ws.merge_cells(start_row=curr_row, start_column=7, end_row=curr_row, end_column=8)
                c_est = ws.cell(row=curr_row, column=7, value=est_val)
                c_est.alignment = Alignment(horizontal="center", vertical="center")
                c_est.font = Font(size=9, color="0F172A", name="Calibri")

                for col_i in range(1, 9):
                    ws.cell(row=curr_row, column=col_i).border = _excel_thin_border()
                    ws.cell(row=curr_row, column=col_i).fill = _excel_fill("FFFFFF")

        curr_row += 1
        ws.row_dimensions[curr_row].height = 4  # espaciador

    curr_row += 1

    # 7. Sección de Resultados (Fila curr_row)
    if hoja_tipo == HOJA_CAIDAS:
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
        sec_res_title = ws.cell(row=curr_row, column=1, value="RESULTADO DE LA INSPECCIÓN")
        sec_res_title.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
        sec_res_title.fill = _excel_fill(EXCEL_HEADER_BG)
        sec_res_title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[curr_row].height = 20
        for c in range(1, 9):
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        curr_row += 1
        ws.row_dimensions[curr_row].height = 24
        es_apto_caida = (getattr(inspeccion, "resultado_general", "apta") == "apta")
        txt_apto = f"  {'[X]' if es_apto_caida else '[  ]'}  El equipo inspeccionado se encuentra apto"
        txt_retirado = f"  {'[X]' if not es_apto_caida else '[  ]'}  El equipo inspeccionado necesita ser retirado de servicio"

        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
        c_apto = ws.cell(row=curr_row, column=1, value=txt_apto)
        c_apto.font = Font(bold=es_apto_caida, size=9.5, color="0F172A", name="Calibri")
        c_apto.fill = _excel_fill("DCFCE7" if es_apto_caida else "FFFFFF")
        c_apto.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 5):
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=8)
        c_ret = ws.cell(row=curr_row, column=5, value=txt_retirado)
        c_ret.font = Font(bold=not es_apto_caida, size=9.5, color="0F172A", name="Calibri")
        c_ret.fill = _excel_fill("FEE2E2" if not es_apto_caida else "FFFFFF")
        c_ret.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(5, 9):
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        curr_row += 1
        ws.row_dimensions[curr_row].height = 22
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        lbl_acc = ws.cell(row=curr_row, column=1, value="Acción tomada:")
        lbl_acc.font = Font(bold=True, size=9, color="334155", name="Calibri")
        lbl_acc.fill = _excel_fill("F8FAFC")
        lbl_acc.alignment = Alignment(horizontal="left", vertical="center")
        for c in [1, 2]:
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=8)
        val_acc = ws.cell(row=curr_row, column=3, value=inspeccion.get_accion_tomada_display() or "—")
        val_acc.font = Font(bold=True, size=9.5, color="0F172A", name="Calibri")
        val_acc.fill = _excel_fill("FFFFFF")
        val_acc.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(3, 9):
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

    else:
        ws.row_dimensions[curr_row].height = 24
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        lbl_res = ws.cell(row=curr_row, column=1, value="Resultado general:")
        lbl_res.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
        lbl_res.fill = _excel_fill(EXCEL_SUBHEADER_BG)
        lbl_res.alignment = Alignment(horizontal="left", vertical="center")
        for c in [1, 2]:
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
        val_res = ws.cell(row=curr_row, column=3, value=inspeccion.get_resultado_general_display() or "No definido")
        res_color_map = {
            "apta": (EXCEL_APTA_BG, EXCEL_APTA_TXT),
            "requiere_reparacion": (EXCEL_REPARACION_BG, EXCEL_REPARACION_TXT),
            "fuera_servicio": (EXCEL_FUERA_BG, EXCEL_FUERA_TXT),
        }
        bg_res, txt_res = res_color_map.get(inspeccion.resultado_general, (EXCEL_SUBHEADER_BG, "FFFFFF"))
        val_res.font = Font(bold=True, size=10, color=txt_res, name="Calibri")
        val_res.fill = _excel_fill(bg_res)
        val_res.alignment = Alignment(horizontal="center", vertical="center")
        for c in [3, 4]:
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
        lbl_acc = ws.cell(row=curr_row, column=5, value="Acción tomada:")
        lbl_acc.font = Font(bold=True, size=9.5, color="FFFFFF", name="Calibri")
        lbl_acc.fill = _excel_fill(EXCEL_SUBHEADER_BG)
        lbl_acc.alignment = Alignment(horizontal="left", vertical="center")
        for c in [5, 6]:
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

        ws.merge_cells(start_row=curr_row, start_column=7, end_row=curr_row, end_column=8)
        val_acc = ws.cell(row=curr_row, column=7, value=inspeccion.get_accion_tomada_display() or "—")
        val_acc.font = Font(bold=True, size=9.5, color="0F172A", name="Calibri")
        val_acc.fill = _excel_fill("F1F5F9")
        val_acc.alignment = Alignment(horizontal="center", vertical="center")
        for c in [7, 8]:
            ws.cell(row=curr_row, column=c).border = _excel_thin_border()

    # Observaciones generales
    curr_row += 1
    ws.row_dimensions[curr_row].height = 22
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
    lbl_obs = ws.cell(row=curr_row, column=1, value="Observaciones:")
    lbl_obs.font = Font(bold=True, size=9.5, color="334155", name="Calibri")
    lbl_obs.fill = _excel_fill("F8FAFC")
    lbl_obs.alignment = Alignment(horizontal="left", vertical="center")
    for c in [1, 2]:
        ws.cell(row=curr_row, column=c).border = _excel_thin_border()

    ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=8)
    val_obs = ws.cell(row=curr_row, column=3, value=inspeccion.observaciones or "Sin observaciones.")
    val_obs.font = Font(bold=False, size=9.5, color="0F172A", name="Calibri")
    val_obs.fill = _excel_fill("FFFFFF")
    val_obs.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(3, 9):
        ws.cell(row=curr_row, column=c).border = _excel_thin_border()

    # 8. Sección de Firmas de Conformidad (4 bloques oficiales)
    curr_row += 2
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
    firma_title = ws.cell(row=curr_row, column=1, value="FIRMAS DE CONFORMIDAD")
    firma_title.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    firma_title.fill = _excel_fill(EXCEL_HEADER_BG)
    firma_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[curr_row].height = 22
    for c in range(1, 9):
        ws.cell(row=curr_row, column=c).border = _excel_thin_border()
    ws.row_dimensions[curr_row].height = 22
    for c in range(1, 9):
        ws.cell(row=curr_row, column=c).border = _excel_thin_border()

    # Espacio para firmas 1 y 2
    for _ in range(3):
        curr_row += 1
        ws.row_dimensions[curr_row].height = 16

    curr_row += 1
    ws.row_dimensions[curr_row].height = 20
    top_line = Border(top=Side(style="medium", color="1E293B"))

    bloques_r1 = [
        (1, 4, "1. Encargado de la actividad"),
        (5, 8, "2. Supervisor del trabajo"),
    ]
    for start_c, end_c, cargo_txt in bloques_r1:
        ws.merge_cells(start_row=curr_row, start_column=start_c, end_row=curr_row, end_column=end_c)
        c_cargo = ws.cell(row=curr_row, column=start_c, value=cargo_txt)
        c_cargo.font = Font(bold=True, size=9.5, color="1E293B", name="Calibri")
        c_cargo.alignment = Alignment(horizontal="center", vertical="center")
        for col_k in range(start_c, end_c + 1):
            ws.cell(row=curr_row, column=col_k).border = top_line

    curr_row += 1
    ws.row_dimensions[curr_row].height = 18
    for start_c, end_c, _ in bloques_r1:
        ws.merge_cells(start_row=curr_row, start_column=start_c, end_row=curr_row, end_column=end_c)
        c_f = ws.cell(row=curr_row, column=start_c, value="Nombre: ______________________  Fecha: ____/____/______")
        c_f.font = Font(bold=False, size=8.5, color="64748B", name="Calibri")
        c_f.alignment = Alignment(horizontal="center", vertical="center")

    # Espacio para firmas 3 y 4
    for _ in range(3):
        curr_row += 1
        ws.row_dimensions[curr_row].height = 16

    curr_row += 1
    ws.row_dimensions[curr_row].height = 20
    bloques_r2 = [
        (1, 4, "3. Responsable del área FM"),
        (5, 8, "4. Responsable de seguridad"),
    ]
    for start_c, end_c, cargo_txt in bloques_r2:
        ws.merge_cells(start_row=curr_row, start_column=start_c, end_row=curr_row, end_column=end_c)
        c_cargo = ws.cell(row=curr_row, column=start_c, value=cargo_txt)
        c_cargo.font = Font(bold=True, size=9.5, color="1E293B", name="Calibri")
        c_cargo.alignment = Alignment(horizontal="center", vertical="center")
        for col_k in range(start_c, end_c + 1):
            ws.cell(row=curr_row, column=col_k).border = top_line

    curr_row += 1
    ws.row_dimensions[curr_row].height = 18
    for start_c, end_c, _ in bloques_r2:
        ws.merge_cells(start_row=curr_row, start_column=start_c, end_row=curr_row, end_column=end_c)
        c_f = ws.cell(row=curr_row, column=start_c, value="Nombre: ______________________  Fecha: ____/____/______")
        c_f.font = Font(bold=False, size=8.5, color="64748B", name="Calibri")
        c_f.alignment = Alignment(horizontal="center", vertical="center")

    widths = [6, 24, 14, 14, 10, 10, 10, 24]
    _excel_set_col_widths(ws, widths)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



# ─── Reporte General Consolidado de Inspecciones ──────────────────────────────

MESES_ES = [
    "", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
]


def generar_excel_inspecciones_generales(almacen_id=None):
    """
    Genera el reporte Excel general y consolidado de Inspecciones para todo el almacén.
    Hojas:
      1. 'Resumen': Conteo total por estado y resultado, KPIs y tablas ejecutivas.
      2. 'Por Mes': Inspecciones realizadas agrupadas por mes/año con resultado.
      3. 'Vencidas': Listado de materiales con inspección vencida ordenado por días de atraso.
      4. 'Top Materiales': Gráfico BarChart de los 15 materiales con más inspecciones no conformes.
    Devuelve (buffer, filename).
    """
    from collections import defaultdict
    from datetime import date
    from django.db.models import Q
    from django.utils import timezone
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Font

    from apps.inspeccion.models import Inspeccion, ProgramacionInspeccion
    from apps.catalogo.models import Almacen

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    hoy = timezone.localdate()

    almacen_nombre = "Todos los almacenes"
    if almacen_id:
        try:
            alm = Almacen.objects.get(pk=almacen_id)
            almacen_nombre = alm.nombre
        except Almacen.DoesNotExist:
            pass

    # ── Consultas base ──
    qs_inspecciones = Inspeccion.objects.select_related(
        "material", "pieza", "inspector", "plantilla", "almacen"
    ).all()
    if almacen_id:
        qs_inspecciones = qs_inspecciones.filter(almacen_id=almacen_id)

    inspecciones = list(qs_inspecciones.order_by("-fecha"))

    # Programaciones vencidas
    qs_vencidas = ProgramacionInspeccion.objects.filter(
        estado="pendiente",
        fecha_programada__lt=hoy,
    ).select_related(
        "material__subcategoria", "pieza__material__subcategoria", "plan"
    )
    if almacen_id:
        qs_vencidas = qs_vencidas.filter(
            Q(almacen_id=almacen_id)
            | Q(material__almacen_id=almacen_id)
            | Q(pieza__material__almacen_id=almacen_id)
        )

    vencidas_list = list(qs_vencidas)
    # Ordenar por días de atraso desc
    vencidas_list.sort(key=lambda p: (hoy - p.fecha_programada).days, reverse=True)

    # ── Métricas calculadas ──
    total_insp = len(inspecciones)
    total_aptas = sum(1 for i in inspecciones if i.resultado_general == "apta")
    total_reparacion = sum(1 for i in inspecciones if i.resultado_general == "requiere_reparacion")
    total_fuera = sum(1 for i in inspecciones if i.resultado_general == "fuera_servicio")
    total_vencidas = len(vencidas_list)

    # ─────────────────────────────────────────────────────────────────────────
    # HOJA 1: Resumen Ejecutivo y KPIs
    # ─────────────────────────────────────────────────────────────────────────
    ws_resumen = wb.create_sheet("Resumen")

    # Título institucional
    ws_resumen.merge_cells("A1:H1")
    t1 = ws_resumen["A1"]
    t1.value = "INCALPACA TOPS S.A. — REPORTE GENERAL DE INSPECCIONES Y CALIDAD"
    t1.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t1.fill = _excel_fill(EXCEL_HEADER_BG)
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[1].height = 30

    # Subtítulo con Almacén y Fecha
    ws_resumen.merge_cells("A2:H2")
    t2 = ws_resumen["A2"]
    t2.value = f"Almacén: {almacen_nombre}  |  Fecha de emisión: {hoy.strftime('%d/%m/%Y')}  |  Total inspecciones registradas: {total_insp}"
    t2.font = Font(bold=False, size=10, color="FFFFFF", name="Calibri")
    t2.fill = _excel_fill(EXCEL_SUBHEADER_BG)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[2].height = 20

    # Tarjetas KPI (Filas 4 a 5)
    kpis = [
        ("A", "B", "TOTAL REALIZADAS", str(total_insp), "1E293B", "F1F5F9", "0F172A"),
        ("C", "C", "APTAS", str(total_aptas), "065F46", "ECFDF5", "065F46"),
        ("D", "E", "REQUIEREN REPARACIÓN", str(total_reparacion), "92400E", "FEF3C7", "92400E"),
        ("F", "G", "FUERA DE SERVICIO", str(total_fuera), "991B1B", "FEE2E2", "991B1B"),
        ("H", "H", "VENCIDAS PENDIENTES", str(total_vencidas), "7F1D1D", "FEE2E2", "991B1B"),
    ]

    ws_resumen.row_dimensions[4].height = 18
    ws_resumen.row_dimensions[5].height = 26

    for start_c, end_c, label, val, bg_head, bg_val, txt_val in kpis:
        c1_idx = ord(start_c) - ord("A") + 1
        c2_idx = ord(end_c) - ord("A") + 1

        ws_resumen.merge_cells(start_row=4, start_column=c1_idx, end_row=4, end_column=c2_idx)
        cell_lbl = ws_resumen.cell(row=4, column=c1_idx, value=label)
        cell_lbl.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        cell_lbl.fill = _excel_fill(bg_head)
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")

        ws_resumen.merge_cells(start_row=5, start_column=c1_idx, end_row=5, end_column=c2_idx)
        cell_val = ws_resumen.cell(row=5, column=c1_idx, value=val)
        cell_val.font = Font(bold=True, size=14, color=txt_val, name="Calibri")
        cell_val.fill = _excel_fill(bg_val)
        cell_val.alignment = Alignment(horizontal="center", vertical="center")

        for r_k in [4, 5]:
            for c_k in range(c1_idx, c2_idx + 1):
                ws_resumen.cell(row=r_k, column=c_k).border = _excel_thin_border()

    # Tabla 1: Distribución por Resultado (Fila 7)
    _excel_write_header_row(ws_resumen, 7, ["Resultado", "Cantidad", "% del Total", "Estado Operativo"])
    res_rows = [
        ("Apta", total_aptas, (f"{total_aptas / total_insp * 100:.1f}%" if total_insp else "0.0%"), "Conforme para uso"),
        ("Requiere reparación", total_reparacion, (f"{total_reparacion / total_insp * 100:.1f}%" if total_insp else "0.0%"), "Observado / En taller"),
        ("Fuera de servicio", total_fuera, (f"{total_fuera / total_insp * 100:.1f}%" if total_insp else "0.0%"), "No apto / Baja requerida"),
    ]
    aligns_t1 = {1: "left", 2: "center", 3: "center", 4: "left"}
    for idx, (res_lbl, cant, pct, est_op) in enumerate(res_rows, start=8):
        c_styles = {}
        if idx == 8:
            c_styles[1] = (EXCEL_APTA_BG, EXCEL_APTA_TXT, True)
        elif idx == 9:
            c_styles[1] = (EXCEL_REPARACION_BG, EXCEL_REPARACION_TXT, True)
        elif idx == 10:
            c_styles[1] = (EXCEL_FUERA_BG, EXCEL_FUERA_TXT, True)
        _excel_write_data_row(ws_resumen, idx, [res_lbl, cant, pct, est_op], is_alt=(idx % 2 == 0), alignments=aligns_t1, cell_styles=c_styles)

    # Tabla 2: Distribución por Inspector (Fila 13)
    _excel_write_header_row(ws_resumen, 13, ["Inspector", "Inspecciones", "Aptas", "Con Observación"])
    insp_stats = defaultdict(lambda: {"total": 0, "aptas": 0, "obs": 0})
    for i in inspecciones:
        nom = i.inspector.get_full_name() or i.inspector.username
        insp_stats[nom]["total"] += 1
        if i.resultado_general == "apta":
            insp_stats[nom]["aptas"] += 1
        else:
            insp_stats[nom]["obs"] += 1

    for idx, (nom_insp, d_insp) in enumerate(sorted(insp_stats.items(), key=lambda x: -x[1]["total"]), start=14):
        _excel_write_data_row(ws_resumen, idx, [
            nom_insp, d_insp["total"], d_insp["aptas"], d_insp["obs"]
        ], is_alt=(idx % 2 == 0), alignments={1: "left", 2: "center", 3: "center", 4: "center"})

    _excel_set_col_widths(ws_resumen, [22, 16, 16, 26, 16, 16, 16, 22])

    # ─────────────────────────────────────────────────────────────────────────
    # HOJA 2: Por Mes
    # ─────────────────────────────────────────────────────────────────────────
    ws_mes = wb.create_sheet("Por Mes")
    cols_mes = ["Año", "Mes", "Aptas", "Requieren Reparación", "Fuera de Servicio", "Total Realizadas", "% Conformidad"]
    _excel_write_header_row(ws_mes, 1, cols_mes)
    _excel_set_col_widths(ws_mes, [8, 10, 14, 22, 18, 18, 16])
    _excel_freeze(ws_mes)

    mes_stats = defaultdict(lambda: {"apta": 0, "reparacion": 0, "fuera": 0, "total": 0})
    for i in inspecciones:
        f = i.fecha.date() if hasattr(i.fecha, "date") else i.fecha
        k = (f.year, f.month)
        mes_stats[k]["total"] += 1
        if i.resultado_general == "apta":
            mes_stats[k]["apta"] += 1
        elif i.resultado_general == "requiere_reparacion":
            mes_stats[k]["reparacion"] += 1
        elif i.resultado_general == "fuera_servicio":
            mes_stats[k]["fuera"] += 1

    rows_mes = sorted(mes_stats.items(), key=lambda x: (x[0][0], x[0][1]))
    aligns_mes = {1: "center", 2: "center", 3: "center", 4: "center", 5: "center", 6: "center", 7: "center"}

    for r_idx, ((anio_k, mes_k), d_m) in enumerate(rows_mes, start=2):
        tot_m = d_m["total"]
        pct_conf = f"{(d_m['apta'] / tot_m * 100):.1f}%" if tot_m else "0.0%"
        _excel_write_data_row(ws_mes, r_idx, [
            anio_k, MESES_ES[mes_k],
            d_m["apta"] or "", d_m["reparacion"] or "", d_m["fuera"] or "", tot_m,
            pct_conf,
        ], is_alt=(r_idx % 2 == 0), alignments=aligns_mes)

    ws_mes.auto_filter.ref = f"A1:G{max(2, len(rows_mes) + 1)}"

    # ─────────────────────────────────────────────────────────────────────────
    # HOJA 3: Vencidas (Materiales con inspección vencida)
    # ─────────────────────────────────────────────────────────────────────────
    ws_vencidas = wb.create_sheet("Vencidas")
    cols_venc = ["Puesto", "Código", "Material / Herramienta", "Subcategoría", "Fecha Programada", "Días de Atraso", "Estado"]
    _excel_write_header_row(ws_vencidas, 1, cols_venc)
    _excel_set_col_widths(ws_vencidas, [8, 14, 38, 22, 18, 16, 14])
    _excel_freeze(ws_vencidas)

    aligns_venc = {1: "center", 2: "center", 3: "left", 4: "left", 5: "center", 6: "center", 7: "center"}

    for idx, prog in enumerate(vencidas_list, start=1):
        r_num = 1 + idx
        dias_atraso = (hoy - prog.fecha_programada).days
        mat = prog.pieza.material if prog.pieza else prog.material
        codigo = prog.pieza.codigo if prog.pieza else (mat.codigo if mat else "—")
        nombre = prog.pieza.nombre if (prog.pieza and prog.pieza.nombre) else (mat.nombre if mat else "—")
        subcat = mat.subcategoria.nombre if (mat and mat.subcategoria) else "—"

        # Celda de días de atraso resaltada en rojo suave
        c_styles = {
            6: ("FEE2E2", "991B1B", True),
            7: (EXCEL_FUERA_BG, "FFFFFF", True),
        }

        _excel_write_data_row(ws_vencidas, r_num, [
            idx,
            codigo,
            nombre,
            subcat,
            prog.fecha_programada.strftime("%d/%m/%Y"),
            f"{dias_atraso} días",
            "VENCIDA",
        ], is_alt=(idx % 2 == 0), alignments=aligns_venc, cell_styles=c_styles)

    ws_vencidas.auto_filter.ref = f"A1:G{max(2, len(vencidas_list) + 1)}"

    # ─────────────────────────────────────────────────────────────────────────
    # HOJA 4: Top Materiales con Más Inspecciones No Conformes + BarChart
    # ─────────────────────────────────────────────────────────────────────────
    ws_top = wb.create_sheet("Top Materiales")

    mat_fallas = defaultdict(lambda: {
        "codigo": "", "nombre": "", "subcategoria": "", "total_fallas": 0, "reparacion": 0, "fuera": 0
    })

    for i in inspecciones:
        if i.resultado_general in ["requiere_reparacion", "fuera_servicio"]:
            mid = i.material_id
            if not mid:
                continue
            st = mat_fallas[mid]
            st["codigo"] = i.material.codigo if i.material else str(mid)
            st["nombre"] = i.material.nombre if i.material else "—"
            st["subcategoria"] = i.material.subcategoria.nombre if (i.material and i.material.subcategoria) else "—"
            st["total_fallas"] += 1
            if i.resultado_general == "requiere_reparacion":
                st["reparacion"] += 1
            elif i.resultado_general == "fuera_servicio":
                st["fuera"] += 1

    sorted_fallas = sorted(mat_fallas.values(), key=lambda x: -x["total_fallas"])[:15]

    # Encabezados de tabla
    cols_top = ["Puesto", "Código", "Material", "Subcategoría", "No Conformes", "Req. Reparación", "Fuera de Servicio"]
    _excel_write_header_row(ws_top, 1, cols_top)
    _excel_set_col_widths(ws_top, [8, 14, 38, 22, 16, 16, 16])
    _excel_freeze(ws_top)

    aligns_top = {1: "center", 2: "center", 3: "left", 4: "left", 5: "center", 6: "center", 7: "center"}

    for idx, mf in enumerate(sorted_fallas, start=1):
        r_num = 1 + idx
        _excel_write_data_row(ws_top, r_num, [
            idx,
            mf["codigo"],
            mf["nombre"],
            mf["subcategoria"],
            mf["total_fallas"],
            mf["reparacion"],
            mf["fuera"],
        ], is_alt=(idx % 2 == 0), alignments=aligns_top)

    ws_top.auto_filter.ref = f"A1:G{max(2, len(sorted_fallas) + 1)}"

    # Insertar Gráfico BarChart si hay datos
    if sorted_fallas:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top 15 Materiales con Más Inspecciones No Conformes"
        chart.y_axis.title = "Cantidad No Conformes"
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.tickLblSkip = 1
        chart.legend = None
        chart.width = 22
        chart.height = 14

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        data = Reference(ws_top, min_col=5, min_row=1, max_row=1 + len(sorted_fallas))
        cats = Reference(ws_top, min_col=3, min_row=2, max_row=1 + len(sorted_fallas))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_top.add_chart(chart, "I2")

    filename = f"reporte_inspecciones_general_{hoy.isoformat()}.xlsx"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, filename

# ─── PDF ───────────────────────────────────────────────────────────────────

TITULOS_PDF = {
    HOJA_MANUALES: "FORMATO DE INSPECCIÓN GRUPAL DE HERRAMIENTAS MANUALES",
    HOJA_INALAMBRICAS: "FORMATO DE INSPECCIÓN DE HERRAMIENTAS ELÉCTRICAS INALÁMBRICAS",
    HOJA_CON_CABLE: "FORMATO DE INSPECCIÓN DE HERRAMIENTAS ELÉCTRICAS CON CABLE",
    HOJA_EPP: "FORMATO DE INSPECCIÓN DE EQUIPO DE PROTECCIÓN PERSONAL (EPP)",
    HOJA_ESCALERAS: "FORMATO DE INSPECCIÓN DE ESCALERAS Y EQUIPOS DE ALTURA",
    HOJA_CAIDAS: "FORMATO DE INSPECCIÓN DE EQUIPOS DE PROTECCIÓN CONTRA CAÍDAS",
}

# Paleta profesional en escala de grises (reemplaza el azul saturado anterior).
GRIS_OSCURO = colors.HexColor("#2b2f36")     # títulos de tabla, líneas fuertes
GRIS_MEDIO = colors.HexColor("#6b7280")      # texto secundario
GRIS_CLARO = colors.HexColor("#f3f4f6")      # zebra striping
GRIS_BORDE = colors.HexColor("#c9ccd1")      # bordes de tabla
NEGRO_TEXTO = colors.HexColor("#1a1c20")

LOGO_PATH = Path(__file__).resolve().parent.parent.parent.parent / "frontend" / "public" / "logo-incalpaca.png"


def generar_pdf_inspeccion(inspeccion):
    hoja_nombre = _detectar_hoja(inspeccion.plantilla.nombre if getattr(inspeccion, "plantilla", None) else "")
    titulo = TITULOS_PDF.get(hoja_nombre, "FORMATO DE INSPECCIÓN DE HERRAMIENTAS")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=0.9 * cm, bottomMargin=0.9 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        "TituloInspeccion", parent=styles["Title"], fontSize=12, leading=14,
        textColor=NEGRO_TEXTO, alignment=TA_CENTER, fontName="Helvetica-Bold",
    )
    subtitulo_style = ParagraphStyle(
        "Subtitulo", parent=styles["Normal"], fontSize=8, textColor=GRIS_MEDIO,
        alignment=TA_CENTER,
    )
    elementos = []

    # ── Encabezado: 3 columnas — Logo Incalpaca FM a la izquierda | Título + Área | Código/Fecha ──
    codigo_doc = _codigo_documento(inspeccion)
    fecha_emision = _fecha_emision_hoy()
    meta_style = ParagraphStyle(
        "MetaDoc", parent=styles["Normal"], fontSize=7.5,
        textColor=GRIS_MEDIO, alignment=TA_RIGHT, leading=10,
    )

    header_logo_path = Path(__file__).resolve().parent / "logo_incalpaca_header.png"
    if not header_logo_path.exists():
        header_logo_path = Path(__file__).resolve().parent / "logo_incalpaca.png"
    if header_logo_path.exists():
        logo_header_widget = Image(str(header_logo_path), width=4.8 * cm, height=0.95 * cm)
    else:
        logo_header_widget = Paragraph("<b>INCALPACA | FM</b>", ParagraphStyle(
            "FMH", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold",
        ))

    # Título central
    titulo_col = Table(
        [[Paragraph(titulo, titulo_style)],
         [Paragraph(f"Área: {getattr(inspeccion, 'area_trabajo', '') or 'Facility Management'}", subtitulo_style)]],
        colWidths=[8.4 * cm],
    )
    titulo_col.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # Columna derecha: Metadatos SST (Código, Versión, Emisión)
    meta_col = Table(
        [[Paragraph(f"Código: <b>{codigo_doc}</b><br/>Versión: <b>01</b><br/>Emisión: <b>{fecha_emision}</b>", meta_style)]],
        colWidths=[4.8 * cm],
    )
    meta_col.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))

    fila_superior = Table(
        [[logo_header_widget, titulo_col, meta_col]],
        colWidths=[5.2 * cm, 8.4 * cm, 4.8 * cm],
    )
    fila_superior.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.8, GRIS_BORDE),
    ]))
    elementos.append(fila_superior)
    elementos.append(Spacer(1, 6))

    # ── Datos generales ──
    inspector_nombre = (
        inspeccion.inspector.get_full_name() or inspeccion.inspector.username
        if inspeccion.inspector else ""
    )
    codigo_objetivo = inspeccion.pieza.codigo if inspeccion.pieza else (inspeccion.material.codigo if inspeccion.material else "-")
    modalidad_txt = "Planificada" if getattr(inspeccion, "modalidad", "planificada") == "planificada" else "No planificada"
    frecuencia_txt = (getattr(inspeccion, "frecuencia", "trimestral") or "trimestral").capitalize()
    area_txt = getattr(inspeccion, "area_trabajo", "") or "Facility Management"
    referencia_ot_txt = getattr(inspeccion, "referencia_orden", "") or "—"

    # Cantidades individuales vs grupales
    if getattr(inspeccion, "tipo", "individual") == "grupal":
        cant_insp_pdf = str(inspeccion.cantidad_inspeccionada if inspeccion.cantidad_inspeccionada is not None else 1)
        cant_apta_num = inspeccion.cantidad_apta if inspeccion.cantidad_apta is not None else 0
        cant_no_apta_num = inspeccion.cantidad_no_apta if inspeccion.cantidad_no_apta is not None else 0
        cant_resumen_pdf = f"{cant_apta_num} aptas / {cant_no_apta_num} no aptas"
    else:
        cant_insp_pdf = "1 (Individual)"
        if inspeccion.cantidad_apta is not None and inspeccion.cantidad_no_apta is not None:
            cant_apta_num = inspeccion.cantidad_apta
            cant_no_apta_num = inspeccion.cantidad_no_apta
        else:
            cant_apta_num = 1 if getattr(inspeccion, "resultado_general", "apta") == "apta" else 0
            cant_no_apta_num = 0 if getattr(inspeccion, "resultado_general", "apta") == "apta" else 1
        cant_resumen_pdf = f"{cant_apta_num} apta{'s' if cant_apta_num != 1 else ''} / {cant_no_apta_num} no apta{'s' if cant_no_apta_num != 1 else ''}"

    from apps.inspeccion.utils import color_inspeccion_actual
    from datetime import date
    fecha_base_pdf = inspeccion.fecha.date() if getattr(inspeccion, "fecha", None) else date.today()
    mes_actual_pdf = fecha_base_pdf.month
    trimestre_activo_num_pdf = ((mes_actual_pdf - 1) // 3) + 1  # 1 a 4

    if hoja_nombre == HOJA_MANUALES and getattr(inspeccion, "tipo", "individual") == "grupal":
        datos = [
            ["Tipo de herramienta / Material:", inspeccion.material.nombre if inspeccion.material else "-", "Responsable:", inspector_nombre],
            ["Tipo de inspección:", modalidad_txt, "Fecha de inspección:", _fecha(inspeccion.fecha)],
            ["Frecuencia planificada:", frecuencia_txt, "Próxima inspección:", _fecha(inspeccion.proxima_inspeccion)],
            ["Área de trabajo / Lugar:", area_txt, "Referencia (OT/OL/OP):", referencia_ot_txt],
            ["Cantidad inspeccionada:", cant_insp_pdf, "Cantidad apta / no apta:", cant_resumen_pdf],
        ]
    else:
        datos = [
            ["Código / Identificador:", codigo_objetivo, "Próxima inspección:", _fecha(inspeccion.proxima_inspeccion)],
            ["Marca / Modelo:", (inspeccion.material.marca if inspeccion.material else "-") or "-", "Tipo de inspección:", modalidad_txt],
            ["Inspector responsable:", inspector_nombre, "Frecuencia planificada:", frecuencia_txt],
            ["Fecha de inspección:", _fecha(inspeccion.fecha), "Área de trabajo / Lugar:", area_txt],
            ["Nombre del equipo / herramienta:", inspeccion.material.nombre if inspeccion.material else "-", "Referencia (OT/OL/OP):", referencia_ot_txt],
        ]

    tabla_datos = Table(datos, colWidths=[4.3 * cm, 4.7 * cm, 4.3 * cm, 4.7 * cm])
    tabla_datos.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), GRIS_OSCURO),
        ("TEXTCOLOR", (2, 0), (2, -1), GRIS_OSCURO),
        ("TEXTCOLOR", (1, 0), (1, -1), NEGRO_TEXTO),
        ("TEXTCOLOR", (3, 0), (3, -1), NEGRO_TEXTO),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE),
        ("BACKGROUND", (0, 0), (0, -1), GRIS_CLARO),
        ("BACKGROUND", (2, 0), (2, -1), GRIS_CLARO),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 5))

    # ── Código de color 5S — leyenda visual compacta con colores oficiales ──
    frecuencia_pdf = (getattr(inspeccion, "frecuencia", "") or "").lower()
    es_bimestral_pdf = (
        frecuencia_pdf == "bimestral"
        or hoja_nombre in [HOJA_INALAMBRICAS, HOJA_CON_CABLE]
    )

    if es_bimestral_pdf:
        periodo_activo_num_pdf = ((mes_actual_pdf - 1) // 2) + 1  # 1 a 6
        _pdf_leyenda_colores = [
            {"q": 1, "nombre": "Amarillo", "meses": "Ene – Feb", "hex": "#FFFF00", "rl": colors.HexColor("#FFFF00"), "txt": colors.HexColor("#000000")},
            {"q": 2, "nombre": "Verde",    "meses": "Mar – Abr", "hex": "#00B050", "rl": colors.HexColor("#00B050"), "txt": colors.white},
            {"q": 3, "nombre": "Rojo",     "meses": "May – Jun", "hex": "#FF0000", "rl": colors.HexColor("#FF0000"), "txt": colors.white},
            {"q": 4, "nombre": "Azul",     "meses": "Jul – Ago", "hex": "#0070C0", "rl": colors.HexColor("#0070C0"), "txt": colors.white},
            {"q": 5, "nombre": "Negro",    "meses": "Sep – Oct", "hex": "#000000", "rl": colors.HexColor("#000000"), "txt": colors.white},
            {"q": 6, "nombre": "Blanco",   "meses": "Nov – Dic", "hex": "#FFFFFF", "rl": colors.HexColor("#FFFFFF"), "txt": colors.HexColor("#000000")},
        ]
        label_color_banner = "<b>Código bimestral de inspección (5S):</b>"
        col_w_inner = 2.3 * cm
        col_w_all = [4.2 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm]
    else:
        periodo_activo_num_pdf = ((mes_actual_pdf - 1) // 3) + 1  # 1 a 4
        _pdf_leyenda_colores = [
            {"q": 1, "nombre": "Amarillo", "meses": "Ene – Mar", "hex": "#FFFF00", "rl": colors.HexColor("#FFFF00"), "txt": colors.HexColor("#000000")},
            {"q": 2, "nombre": "Verde",    "meses": "Abr – Jun", "hex": "#00B050", "rl": colors.HexColor("#00B050"), "txt": colors.white},
            {"q": 3, "nombre": "Azul",     "meses": "Jul – Sep", "hex": "#0070C0", "rl": colors.HexColor("#0070C0"), "txt": colors.white},
            {"q": 4, "nombre": "Rojo",     "meses": "Oct – Dic", "hex": "#FF0000", "rl": colors.HexColor("#FF0000"), "txt": colors.white},
        ]
        label_color_banner = "<b>Código trimestral de inspección (5S):</b>"
        col_w_inner = 3.2 * cm
        col_w_all = [5.2 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm]

    _color_badge_style = ParagraphStyle(
        "ColorBadge", parent=styles["Normal"], fontSize=6.5 if es_bimestral_pdf else 7, leading=8,
        fontName="Helvetica-Bold", alignment=1,
    )
    _color_sub_style = ParagraphStyle(
        "ColorSub", parent=styles["Normal"], fontSize=5.5 if es_bimestral_pdf else 6, leading=7,
        fontName="Helvetica", alignment=1,
    )
    _color_label_style = ParagraphStyle(
        "ColorLabel", parent=styles["Normal"], fontSize=7.5, leading=9,
        fontName="Helvetica-Bold", textColor=GRIS_OSCURO,
    )

    color_row = [Paragraph(label_color_banner, _color_label_style)]
    for litem in _pdf_leyenda_colores:
        is_active = (litem["q"] == periodo_activo_num_pdf)
        pref = "B" if es_bimestral_pdf else "Q"
        badge_text = f"<b>{pref}{litem['q']}: {litem['nombre']}{' (VIG)' if is_active else ''}</b>"
        sub_text = litem["meses"]
        bg_color = litem["rl"]
        txt_color = litem["txt"]
        p_badge = Paragraph(badge_text, ParagraphStyle("CB", parent=_color_badge_style, textColor=txt_color))
        p_sub = Paragraph(sub_text, ParagraphStyle("CS", parent=_color_sub_style, textColor=txt_color))
        inner = Table([[p_badge], [p_sub]], colWidths=[col_w_inner])
        if is_active:
            inner.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg_color),
                ("BOX", (0, 0), (-1, -1), 1.8, colors.HexColor("#0F172A")),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
        else:
            inner.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg_color),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
        color_row.append(inner)

    color_banner = Table(
        [color_row],
        colWidths=col_w_all,
    )
    color_banner.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE),
        ("BACKGROUND", (0, 0), (0, 0), GRIS_CLARO),
    ]))
    elementos.append(color_banner)
    elementos.append(Spacer(1, 5))

    # ── Tipo de herramientas manuales — SOLO si es manual y el usuario seleccionó al menos un tipo ──
    tipos_marcados = getattr(inspeccion, "tipos_herramientas", []) or []
    if hoja_nombre == HOJA_MANUALES and len(tipos_marcados) > 0:
        def _chk(nombre_grupo):
            return "[X]" if nombre_grupo in tipos_marcados else "[  ]"

        tabla_tipos_h = Table(
            [
                [Paragraph("<b>TIPO DE HERRAMIENTAS MANUALES</b>", styles["Normal"]), ""],
                [
                    Paragraph(f"{_chk('Herramientas de golpe')} Herramientas de golpe &nbsp;&nbsp;&nbsp; {_chk('Herramientas de corte')} Herramientas de corte &nbsp;&nbsp;&nbsp; {_chk('Herramientas de cohesión')} Herramientas de cohesión &nbsp;&nbsp;&nbsp; {_chk('Herramientas de torsión y ajuste')} Torsión y ajuste", styles["Normal"]),
                    ""
                ],
                [
                    Paragraph(f"{_chk('Herramientas de medición')} Herramientas de medición &nbsp;&nbsp;&nbsp; {_chk('Herramientas de sujeción')} Herramientas de sujeción &nbsp;&nbsp;&nbsp; {_chk('Herramientas de pintura')} Herramientas de pintura &nbsp;&nbsp;&nbsp; {_chk('Otras herramientas')} Otras herramientas", styles["Normal"]),
                    ""
                ],
            ],
            colWidths=[18 * cm, 0 * cm],
        )
        tabla_tipos_h.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GRIS_OSCURO),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (0, -1), 0.4, GRIS_BORDE),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elementos.append(tabla_tipos_h)
        elementos.append(Spacer(1, 5))

    # ── Criterios ──
    criterio_texto_style = ParagraphStyle(
        "CriterioTexto", parent=styles["Normal"], fontSize=8, leading=9.6,
    )
    criterio_obs_style = ParagraphStyle(
        "CriterioObs", parent=styles["Normal"], fontSize=7.5, leading=9,
    )

    plantilla_obj_pdf = (
        inspeccion.plantilla
        if getattr(inspeccion, "plantilla", None)
        else (inspeccion.material.subcategoria.plantilla_inspeccion if getattr(inspeccion, "material", None) and getattr(inspeccion.material, "subcategoria", None) else None)
    )
    if not plantilla_obj_pdf:
        from apps.inspeccion.models import PlantillaCriterio
        nombre_buscar = "Inalámbrica" if hoja_nombre == HOJA_INALAMBRICAS else ("Eléctrica" if hoja_nombre == HOJA_CON_CABLE else ("Manual" if hoja_nombre == HOJA_MANUALES else ("EPP (equipo de protección personal)" if hoja_nombre == HOJA_EPP else ("Escaleras" if hoja_nombre == HOJA_ESCALERAS else "Equipos de protección contra caídas"))))
        plantilla_obj_pdf = PlantillaCriterio.objects.filter(nombre__icontains=nombre_buscar.split()[0]).first()

    criterios_oficiales_pdf = list(plantilla_obj_pdf.criterios.all().order_by("orden")) if plantilla_obj_pdf else []
    respuestas_dict_pdf = {r.criterio_id: r for r in inspeccion.respuestas.select_related("criterio").all()}

    filas_criterios_pdf = []
    if criterios_oficiales_pdf:
        for crit in criterios_oficiales_pdf:
            resp = respuestas_dict_pdf.get(crit.id)
            valor_resp = resp.valor if resp else ("cumple" if getattr(inspeccion, "resultado_general", "apta") == "apta" else "")
            obs_resp = resp.observacion if resp else ""
            filas_criterios_pdf.append((crit.orden, crit.texto, valor_resp, obs_resp))
    elif respuestas_dict_pdf:
        for r in sorted(respuestas_dict_pdf.values(), key=lambda x: x.criterio.orden):
            filas_criterios_pdf.append((r.criterio.orden, r.criterio.texto, r.valor, r.observacion or ""))

    if hoja_nombre == HOJA_CAIDAS:
        data = [["N°", "Criterio de inspección (Defectos a evaluar)", "SÍ", "NO", "No aplica", "Obs."]]
        for crit_ord, crit_txt, resp_val, resp_obs in filas_criterios_pdf:
            data.append([
                crit_ord,
                Paragraph(crit_txt, criterio_texto_style),
                "X" if resp_val == "no_cumple" else "",
                "X" if resp_val == "cumple" else "",
                "X" if resp_val == "no_aplica" else "",
                Paragraph(resp_obs, criterio_obs_style) if resp_obs else "",
            ])
    else:
        data = [["N°", "Criterio de inspección", "Cumple", "No cumple", "No aplica", "Obs."]]
        for crit_ord, crit_txt, resp_val, resp_obs in filas_criterios_pdf:
            data.append([
                crit_ord,
                Paragraph(crit_txt, criterio_texto_style),
                "X" if resp_val == "cumple" else "",
                "X" if resp_val == "no_cumple" else "",
                "X" if resp_val == "no_aplica" else "",
                Paragraph(resp_obs, criterio_obs_style) if resp_obs else "",
            ])

    tabla = Table(data, colWidths=[1.2 * cm, 7.3 * cm, 1.9 * cm, 2.1 * cm, 1.9 * cm, 3.6 * cm], repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GRIS_OSCURO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (4, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 6))

    seccion_heading_style = ParagraphStyle(
        "SeccionHeading", parent=styles["Normal"], fontSize=8.5, fontName="Helvetica-Bold",
        textColor=GRIS_OSCURO, spaceBefore=0, spaceAfter=2,
    )
    seccion_body_style = ParagraphStyle(
        "SeccionBody", parent=styles["Normal"], fontSize=8, leading=12,
    )

    # ── Tabla de Herramientas / EPP con Observaciones — SOLO para Manuales y EPP ──
    if hoja_nombre == HOJA_EPP:
        obs_items = list(inspeccion.items_con_observacion.all())
        obs_data = [["Código", "Nombre del EPP", "Observación encontrada"]]
        filas = obs_items if len(obs_items) > 0 else [None]
        for item in filas:
            obs_data.append([
                item.codigo if item else "-",
                Paragraph(item.nombre, styles["Normal"]) if item else "-",
                Paragraph(item.observacion_encontrada, styles["Normal"]) if item else "-",
            ])
        t_obs = Table(obs_data, colWidths=[3.5 * cm, 5.5 * cm, 9.0 * cm])
        t_obs.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GRIS_OSCURO),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elementos.append(Paragraph("<b>EPP CON OBSERVACIONES (condición insegura)</b>", seccion_heading_style))
        elementos.append(t_obs)
        elementos.append(Spacer(1, 5))
    elif hoja_nombre == HOJA_MANUALES:
        obs_items = list(inspeccion.items_con_observacion.all())
        obs_data = [["Código", "Nombre de la herramienta", "Observación encontrada", "Acción recomendada", "Estado"]]
        filas = obs_items if len(obs_items) > 0 else [None]
        for item in filas:
            obs_data.append([
                item.codigo if item else "-",
                Paragraph(item.nombre, styles["Normal"]) if item else "-",
                Paragraph(item.observacion_encontrada, styles["Normal"]) if item else "-",
                Paragraph(item.accion_recomendada or "-", styles["Normal"]) if item else "-",
                item.estado or "-" if item else "-",
            ])
        t_obs = Table(obs_data, colWidths=[2.2 * cm, 4.0 * cm, 5.0 * cm, 4.5 * cm, 2.3 * cm])
        t_obs.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GRIS_OSCURO),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elementos.append(Paragraph("<b>HERRAMIENTAS CON OBSERVACIONES (condición insegura)</b>", seccion_heading_style))
        elementos.append(t_obs)
        elementos.append(Spacer(1, 5))

    # ── Resultado y acción ──
    def _marca(activo):
        return "[X]" if activo else "[  ]"

    if hoja_nombre == HOJA_CAIDAS:
        es_apto_pdf = (inspeccion.resultado_general == 'apta')
        resultado_txt = (
            f"{_marca(es_apto_pdf)} El equipo inspeccionado se encuentra apto &nbsp;&nbsp;&nbsp;&nbsp;"
            f"{_marca(not es_apto_pdf)} El equipo inspeccionado necesita ser retirado de servicio"
        )
        elementos.append(Paragraph("RESULTADO DE LA INSPECCIÓN", seccion_heading_style))
        elementos.append(Paragraph(resultado_txt, seccion_body_style))
        if getattr(inspeccion, "accion_tomada", None):
            elementos.append(Spacer(1, 3))
            elementos.append(Paragraph("ACCIÓN TOMADA", seccion_heading_style))
            accion_txt = (
                f"{_marca(inspeccion.accion_tomada == 'continua_servicio')} Continúa en servicio &nbsp;&nbsp;"
                f"{_marca(inspeccion.accion_tomada == 'enviar_reparacion')} Enviar a reparación &nbsp;&nbsp;"
                f"{_marca(inspeccion.accion_tomada == 'retirar_servicio')} Retirar del servicio &nbsp;&nbsp;"
                f"{_marca(inspeccion.accion_tomada == 'dar_baja')} Dar de baja &nbsp;&nbsp;"
                f"{_marca(inspeccion.accion_tomada == 'reemplazar')} Reemplazar"
            )
            elementos.append(Paragraph(accion_txt, seccion_body_style))
    elif hoja_nombre != HOJA_MANUALES:
        resultado_txt = (
            f"{_marca(inspeccion.resultado_general == 'apta')} Apta &nbsp;&nbsp;&nbsp;"
            f"{_marca(inspeccion.resultado_general == 'requiere_reparacion')} Requiere reparación &nbsp;&nbsp;&nbsp;"
            f"{_marca(inspeccion.resultado_general == 'fuera_servicio')} Fuera de servicio"
        )
        accion_txt = (
            f"{_marca(inspeccion.accion_tomada == 'continua_servicio')} Continúa en servicio &nbsp;&nbsp;"
            f"{_marca(inspeccion.accion_tomada == 'enviar_reparacion')} Enviar a reparación &nbsp;&nbsp;"
            f"{_marca(inspeccion.accion_tomada == 'retirar_servicio')} Retirar del servicio &nbsp;&nbsp;"
            f"{_marca(inspeccion.accion_tomada == 'dar_baja')} Dar de baja &nbsp;&nbsp;"
            f"{_marca(inspeccion.accion_tomada == 'reemplazar')} Reemplazar"
        )
        elementos.append(Paragraph("RESULTADO DE LA INSPECCIÓN", seccion_heading_style))
        elementos.append(Paragraph(resultado_txt, seccion_body_style))
        elementos.append(Spacer(1, 3))
        elementos.append(Paragraph("ACCIÓN TOMADA", seccion_heading_style))
        elementos.append(Paragraph(accion_txt, seccion_body_style))
    else:
        aptas_txt = (
            f"{_marca(not inspeccion.cantidad_no_apta)} Todas las herramientas inspeccionadas se encuentran aptas.<br/>"
            f"{_marca(bool(inspeccion.cantidad_no_apta))} Existen herramientas con observaciones."
        )
        elementos.append(Paragraph("RESULTADO FINAL", seccion_heading_style))
        elementos.append(Paragraph(aptas_txt, seccion_body_style))

    obs_heading_style = ParagraphStyle(
        "ObsHeading", parent=styles["Heading4"], fontSize=9, textColor=GRIS_OSCURO,
        spaceBefore=0, spaceAfter=2,
    )
    obs_style = ParagraphStyle("ObsBody", parent=styles["Normal"], fontSize=8, leading=11)
    elementos.append(Spacer(1, 4))
    elementos.append(Paragraph("OBSERVACIONES GENERALES", obs_heading_style))
    elementos.append(Paragraph(inspeccion.observaciones or "-", obs_style))
    elementos.append(Spacer(1, 3))

    # ── Firmas (más espacio en blanco para firmar a mano, letras reducidas) ──
    # Se agrupa en KeepTogether para que la fila de nombres/roles nunca quede
    # separada del resto del bloque en un salto de página.
    firma_label_style = ParagraphStyle(
        "FirmaLabel", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Bold",
        textColor=NEGRO_TEXTO, alignment=TA_CENTER,
    )
    firma_dato_style = ParagraphStyle(
        "FirmaDato", parent=styles["Normal"], fontSize=7.5, textColor=GRIS_MEDIO,
        alignment=TA_CENTER, leading=11,
    )
    firmas = Table(
        [
            ["", "", ""],  # espacio amplio (5 renglones) para la firma manuscrita y sello
            [Paragraph("Inspector", firma_label_style),
             Paragraph("Supervisor SST / Mantenimiento", firma_label_style),
             Paragraph("Responsable del Área", firma_label_style)],
            [Paragraph("Nombre: _____________________", firma_dato_style),
             Paragraph("Nombre: _____________________", firma_dato_style),
             Paragraph("Nombre: _____________________", firma_dato_style)],
            [Paragraph("Fecha: ____ / ____ / ______", firma_dato_style),
             Paragraph("Fecha: ____ / ____ / ______", firma_dato_style),
             Paragraph("Fecha: ____ / ____ / ______", firma_dato_style)],
        ],
        colWidths=[6 * cm, 6 * cm, 6 * cm],
        rowHeights=[1.8 * cm, None, None, None],
    )
    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("LINEABOVE", (0, 1), (-1, 1), 0.8, GRIS_OSCURO),
        ("TOPPADDING", (0, 1), (-1, 1), 3),
        ("TOPPADDING", (0, 2), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 0),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
    ]))
    elementos.append(KeepTogether(firmas))

    doc.build(elementos)
    buffer.seek(0)
    return buffer

# ─── HISTORIAL DE INSPECCIONES POR MATERIAL ─────────────────────────────────
# A diferencia de generar_excel_inspeccion/generar_pdf_inspeccion (que llenan
# el formato oficial de UNA inspección puntual), estas dos funciones arman un
# reporte tabular con TODAS las inspecciones históricas de un material —
# pensado para auditoría/trazabilidad, no para el formato SST oficial.

RESULTADO_LABELS_HISTORIAL = {
    "apta": "Apta",
    "requiere_reparacion": "Requiere reparación",
    "fuera_servicio": "Fuera de servicio",
}
ACCION_LABELS_HISTORIAL = {
    "continua_servicio": "Continúa en servicio",
    "enviar_reparacion": "Enviar a reparación",
    "retirar_servicio": "Retirar del servicio",
    "dar_baja": "Dar de baja",
    "reemplazar": "Reemplazar",
}

def _filas_historial_material(material):
    """Devuelve las inspecciones del material, ordenadas de más reciente a más
    antigua, junto con los valores ya formateados para las columnas del reporte."""
    # Ordenamos de forma cronológica ASCENDENTE primero, solo para poder asignar
    # el número secuencial (#1, #2, #3...) según el orden real en que se hicieron
    # las inspecciones DE ESTE MATERIAL. Usar insp.id aquí sería incorrecto porque
    # el id es un correlativo global de toda la tabla de inspecciones (compartido
    # entre todos los materiales), no un contador propio del material.
    inspecciones = list(
        material.inspecciones
        .select_related("pieza", "inspector")
        .order_by("fecha", "id")
    )
    filas = []
    for idx, insp in enumerate(inspecciones, start=1):
        inspector_nombre = (
            (insp.inspector.get_full_name() or insp.inspector.username)
            if insp.inspector else "—"
        )
        filas.append({
            "fecha": _fecha(insp.fecha),
            "numero": f"#{idx}",
            "tipo": "Individual" if insp.tipo == "individual" else "Grupal",
            "pieza": insp.pieza.codigo if insp.pieza else "—",
            "responsable": inspector_nombre,
            "resultado": RESULTADO_LABELS_HISTORIAL.get(insp.resultado_general, insp.resultado_general or "—"),
            "accion": ACCION_LABELS_HISTORIAL.get(insp.accion_tomada, insp.accion_tomada or "—"),
            "proxima": _fecha(insp.proxima_inspeccion),
            "observaciones": insp.observaciones or "",
        })
    # El reporte se muestra de más reciente a más antigua (igual que antes,
    # que ordenaba por "-fecha"); como ya numeramos en orden ascendente,
    # simplemente invertimos la lista para la presentación.
    filas.reverse()
    return filas

def generar_excel_historial_material(material):
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    filas = _filas_historial_material(material)
    total = len(filas)
    n_apta = sum(1 for f in filas if f["resultado"] == "Apta")
    n_reparacion = sum(1 for f in filas if f["resultado"] == "Requiere reparación")
    n_fuera = sum(1 for f in filas if f["resultado"] == "Fuera de servicio")

    GRIS_OSCURO_HEX = "2B2F36"
    GRIS_CLARO_HEX = "F3F4F6"
    GRIS_MEDIO_HEX = "6B7280"
    NEGRO_HEX = "1A1C20"

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de inspecciones"

    borde = Border(*[Side(style="thin", color="C9CCD1")] * 4)

    # ── Marca (logo + nombre de empresa) ──
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width = 40
            img.height = 40
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.merge_cells("C1:I1")
    ws["C1"] = "INCALPACA"
    ws["C1"].font = Font(bold=True, size=14, color=NEGRO_HEX)
    ws.merge_cells("C2:I2")
    ws["C2"] = "Facility Management · Sistema de Gestión de Almacén"
    ws["C2"].font = Font(size=8.5, italic=True, color=GRIS_MEDIO_HEX)

    # ── Título del reporte ──
    fila = 4
    ws.merge_cells(f"A{fila}:I{fila}")
    ws[f"A{fila}"] = f"Historial de inspecciones — {material.codigo} · {material.nombre}"
    ws[f"A{fila}"].font = Font(bold=True, size=13, color=NEGRO_HEX)
    ws[f"A{fila}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[fila].height = 22

    fila += 1
    categoria_nombre = getattr(material, "categoria_nombre", None)
    if not categoria_nombre:
        subcategoria = getattr(material, "subcategoria", None)
        categoria = getattr(subcategoria, "categoria", None)
        categoria_nombre = getattr(categoria, "nombre", None)
    ws.merge_cells(f"A{fila}:I{fila}")
    ws[f"A{fila}"] = (
        f"Marca: {material.marca or '—'}   ·   Categoría: {categoria_nombre or '—'}   ·   "
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws[f"A{fila}"].font = Font(italic=True, size=9, color=GRIS_MEDIO_HEX)
    ws[f"A{fila}"].alignment = Alignment(horizontal="center")

    # ── KPIs de resumen ──
    fila += 2
    fila_kpi = fila
    kpis = [
        ("Total inspecciones", total, GRIS_OSCURO_HEX, "FFFFFF"),
        ("Aptas", n_apta, "DCFCE7", "15803D"),
        ("Requiere reparación", n_reparacion, "FEF3C7", "B45309"),
        ("Fuera de servicio", n_fuera, "FEE2E2", "B91C1C"),
    ]
    col = 1
    for etiqueta, valor, bg, fg in kpis:
        c1, c2 = get_column_letter(col), get_column_letter(col + 1)
        ws.merge_cells(f"{c1}{fila_kpi}:{c2}{fila_kpi}")
        celda = ws[f"{c1}{fila_kpi}"]
        celda.value = f"{etiqueta}: {valor}"
        celda.font = Font(bold=True, size=9.5, color=fg)
        celda.fill = PatternFill("solid", fgColor=bg)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        for c in (c1, c2):
            ws[f"{c}{fila_kpi}"].border = borde
        col += 2
    ws.row_dimensions[fila_kpi].height = 20

    # ── Tabla de historial ──
    fila_encabezado = fila_kpi + 2
    encabezados = [
        "Fecha", "N° Inspección", "Tipo", "Código de pieza", "Responsable",
        "Resultado", "Acción tomada", "Próxima inspección", "Observaciones",
    ]
    fill_encabezado = PatternFill("solid", fgColor=GRIS_OSCURO_HEX)
    fuente_encabezado = Font(bold=True, color="FFFFFF", size=9.5)
    for col_i, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col_i, value=titulo)
        celda.fill = fill_encabezado
        celda.font = fuente_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde
    ws.row_dimensions[fila_encabezado].height = 26

    fill_resultado = {
        "Apta": PatternFill("solid", fgColor="DCFCE7"),
        "Requiere reparación": PatternFill("solid", fgColor="FEF3C7"),
        "Fuera de servicio": PatternFill("solid", fgColor="FEE2E2"),
    }
    color_resultado = {
        "Apta": "15803D",
        "Requiere reparación": "B45309",
        "Fuera de servicio": "B91C1C",
    }
    fila_zebra = PatternFill("solid", fgColor=GRIS_CLARO_HEX)

    for i, item in enumerate(filas):
        f = fila_encabezado + 1 + i
        valores = [
            item["fecha"], item["numero"], item["tipo"], item["pieza"],
            item["responsable"], item["resultado"], item["accion"],
            item["proxima"], item["observaciones"],
        ]
        for col_i, valor in enumerate(valores, start=1):
            celda = ws.cell(row=f, column=col_i, value=valor)
            celda.border = borde
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2 == 1:
                celda.fill = fila_zebra
        celda_resultado = ws.cell(row=f, column=6)
        if item["resultado"] in fill_resultado:
            celda_resultado.fill = fill_resultado[item["resultado"]]
            celda_resultado.font = Font(bold=True, color=color_resultado[item["resultado"]])

    anchos = [12, 12, 10, 14, 20, 18, 20, 16, 38]
    for col_i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = ancho

    ultima_fila = max(fila_encabezado, fila_encabezado + len(filas))
    ws.auto_filter.ref = f"A{fila_encabezado}:I{ultima_fila}"
    ws.freeze_panes = f"A{fila_encabezado + 1}"

    # Impresión: apaisado, ajustado al ancho, repitiendo la fila de encabezado
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{fila_encabezado}:{fila_encabezado}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_historial_material(material):
    from datetime import datetime

    filas = _filas_historial_material(material)
    total = len(filas)
    n_apta = sum(1 for f in filas if f["resultado"] == "Apta")
    n_reparacion = sum(1 for f in filas if f["resultado"] == "Requiere reparación")
    n_fuera = sum(1 for f in filas if f["resultado"] == "Fuera de servicio")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        topMargin=1 * cm, bottomMargin=1.3 * cm,
        leftMargin=1.3 * cm, rightMargin=1.3 * cm,
    )
    styles = getSampleStyleSheet()

    empresa_style = ParagraphStyle(
        "EmpresaHist", parent=styles["Normal"], fontSize=11.5,
        textColor=NEGRO_TEXTO, fontName="Helvetica-Bold", leading=13,
    )
    meta_style = ParagraphStyle(
        "MetaHist", parent=styles["Normal"], fontSize=7.5,
        textColor=GRIS_MEDIO, alignment=TA_RIGHT, leading=10,
    )
    titulo_style = ParagraphStyle(
        "TituloHistorial", parent=styles["Title"], fontSize=14, leading=17,
        textColor=NEGRO_TEXTO, alignment=TA_CENTER, fontName="Helvetica-Bold",
    )
    subtitulo_style = ParagraphStyle(
        "SubtituloHistorial", parent=styles["Normal"], fontSize=9, textColor=GRIS_MEDIO,
        alignment=TA_CENTER,
    )
    celda_style = ParagraphStyle("CeldaHistorial", parent=styles["Normal"], fontSize=8, leading=9.5)

    elementos = []

    # ── Marca (igual que el formato de inspección individual) ──
    if LOGO_PATH.exists():
        logo_img = Image(str(LOGO_PATH), width=1.1 * cm, height=1.1 * cm)
    else:
        logo_img = Paragraph("", styles["Normal"])
    marca_cell = Table(
        [[logo_img, Paragraph("INCALPACA<br/><font size=6.5 color='#6b7280'>Facility Management</font>", empresa_style)]],
        colWidths=[1.35 * cm, 6 * cm],
    )
    marca_cell.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    meta_cell = Paragraph(
        f"Generado el: <b>{datetime.now().strftime('%d/%m/%Y %H:%M')}</b><br/>Total de registros: <b>{total}</b>",
        meta_style,
    )
    fila_superior = Table([[marca_cell, meta_cell]], colWidths=[16 * cm, 8 * cm])
    fila_superior.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elementos.append(fila_superior)
    elementos.append(Spacer(1, 6))

    elementos.append(Paragraph("HISTORIAL DE INSPECCIONES", titulo_style))
    elementos.append(Spacer(1, 2))
    elementos.append(Paragraph(f"{material.codigo} · {material.nombre}", subtitulo_style))
    elementos.append(Spacer(1, 8))

    # ── KPIs de resumen ──
    kpi_style = ParagraphStyle("KpiHist", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", alignment=TA_CENTER)

    def kpi_cell(texto, color_texto):
        estilo = ParagraphStyle("KpiHistColor", parent=kpi_style, textColor=color_texto)
        return Paragraph(texto, estilo)

    kpis_data = [[
        kpi_cell(f"Total<br/>{total}", NEGRO_TEXTO),
        kpi_cell(f"Aptas<br/>{n_apta}", colors.HexColor("#15803d")),
        kpi_cell(f"Requiere reparación<br/>{n_reparacion}", colors.HexColor("#b45309")),
        kpi_cell(f"Fuera de servicio<br/>{n_fuera}", colors.HexColor("#b91c1c")),
    ]]
    kpi_bgs = [GRIS_CLARO, colors.HexColor("#dcfce7"), colors.HexColor("#fef3c7"), colors.HexColor("#fee2e2")]
    kpis_tabla = Table(kpis_data, colWidths=[6 * cm] * 4, rowHeights=[1.1 * cm])
    estilo_kpi = [("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE)]
    for idx, bg in enumerate(kpi_bgs):
        estilo_kpi.append(("BACKGROUND", (idx, 0), (idx, 0), bg))
    kpis_tabla.setStyle(TableStyle(estilo_kpi))
    elementos.append(kpis_tabla)
    elementos.append(Spacer(1, 10))

    # ── Tabla de historial ──
    encabezados = ["Fecha", "N°", "Tipo", "Código de pieza", "Responsable", "Resultado", "Acción tomada", "Próx. inspección", "Observaciones"]
    data = [encabezados]
    resultado_bg = {
        "Apta": colors.HexColor("#dcfce7"),
        "Requiere reparación": colors.HexColor("#fef3c7"),
        "Fuera de servicio": colors.HexColor("#fee2e2"),
    }
    resultado_fg = {
        "Apta": colors.HexColor("#15803d"),
        "Requiere reparación": colors.HexColor("#b45309"),
        "Fuera de servicio": colors.HexColor("#b91c1c"),
    }
    for item in filas:
        resultado_style = ParagraphStyle(
            "ResultadoHist", parent=celda_style, fontName="Helvetica-Bold",
            textColor=resultado_fg.get(item["resultado"], NEGRO_TEXTO),
        )
        data.append([
            item["fecha"], item["numero"], item["tipo"], item["pieza"],
            Paragraph(item["responsable"], celda_style),
            Paragraph(item["resultado"], resultado_style),
            Paragraph(item["accion"], celda_style),
            item["proxima"],
            Paragraph(item["observaciones"], celda_style),
        ])

    tabla = Table(
        data,
        colWidths=[2.1 * cm, 1.4 * cm, 1.8 * cm, 2.6 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 2.4 * cm, 4.8 * cm],
        repeatRows=1,
    )
    estilo_tabla = [
        ("BACKGROUND", (0, 0), (-1, 0), GRIS_OSCURO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (3, -1), "CENTER"),
        ("ALIGN", (7, 0), (7, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, GRIS_BORDE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, item in enumerate(filas, start=1):
        if item["resultado"] in resultado_bg:
            estilo_tabla.append(("BACKGROUND", (5, i), (5, i), resultado_bg[item["resultado"]]))
        elif i % 2 == 0:
            estilo_tabla.append(("BACKGROUND", (0, i), (-1, i), GRIS_CLARO))
    tabla.setStyle(TableStyle(estilo_tabla))
    elementos.append(tabla)

    if not filas:
        elementos.append(Spacer(1, 12))
        elementos.append(Paragraph("Este material no tiene inspecciones registradas todavía.", subtitulo_style))

    def pie_pagina(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(GRIS_MEDIO)
        ancho, _ = landscape(letter)
        canvas.drawString(1.3 * cm, 0.7 * cm, f"INCALPACA · Historial de inspecciones — {material.codigo}")
        canvas.drawRightString(ancho - 1.3 * cm, 0.7 * cm, f"Página {doc_.page}")
        canvas.restoreState()

    doc.build(elementos, onFirstPage=pie_pagina, onLaterPages=pie_pagina)
    buffer.seek(0)
    return buffer