from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_spectacular.utils import extend_schema, inline_serializer
from rest_framework import generics, permissions, response, serializers, status, views
from openpyxl import load_workbook
from django.db import transaction
from rest_framework_simplejwt.views import TokenRefreshView
from django.contrib.auth import get_user_model

from .serializers import ChangePasswordSerializer, CurrentUserSerializer, LoginSerializer, UserListSerializer, TechnicianSerializer
from .models import AccountProfile
from .permissions import IsAdministrator
from apps.notifications.services import queue_notification
from config.schema import DetailResponseSerializer, ImportResultSerializer


LoginResponseSerializer = inline_serializer(
    name="LoginResponse",
    fields={
        "access": serializers.CharField(),
        "refresh": serializers.CharField(),
        "user": CurrentUserSerializer,
    },
)
TechnicianImportSerializer = inline_serializer(
    name="TechnicianImportRequest", fields={"file": serializers.FileField()}
)
TechnicianNotificationSerializer = inline_serializer(
    name="TechnicianNotificationRequest",
    fields={
        "template": serializers.CharField(required=False),
        "deliveryChannel": serializers.CharField(required=False),
        "subject": serializers.CharField(required=False),
        "body": serializers.CharField(required=False),
    },
)

class LoginView(views.APIView):
    permission_classes = [permissions.AllowAny]
    throttle_scope = "login"

    @extend_schema(request=LoginSerializer, responses={200: LoginResponseSerializer})
    def post(self, request):
        serializer = LoginSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        return response.Response(serializer.validated_data)


class CurrentUserView(views.APIView):
    @extend_schema(responses={200: CurrentUserSerializer})
    def get(self, request):
        return response.Response(CurrentUserSerializer(request.user).data)


class ChangePasswordView(views.APIView):
    @extend_schema(request=ChangePasswordSerializer, responses={200: DetailResponseSerializer})
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return response.Response(
            {"detail": "Contraseña actualizada correctamente."}, status=status.HTTP_200_OK
        )


class UserListView(views.APIView):
    """Lista todos los usuarios activos. Temporal con AllowAny hasta que exista autenticación en el frontend."""
    permission_classes = [permissions.AllowAny]

    @extend_schema(responses={200: UserListSerializer(many=True)})
    def get(self, request):
        User = get_user_model()
        users = (
            User.objects
            .filter(is_active=True)
            .select_related("account_profile")
            .order_by("first_name", "last_name")
        )
        serializer = UserListSerializer(users, many=True)
        return response.Response(serializer.data)


class TechnicianListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAdministrator]
    serializer_class = TechnicianSerializer

    def get_queryset(self):
        return get_user_model().objects.select_related('account_profile').filter(
            account_profile__role__in=[
                AccountProfile.Role.TECHNICIAN,
                AccountProfile.Role.ALMACENERO,
            ]
        ).order_by('first_name', 'last_name', 'username')


class TechnicianImportView(views.APIView):
    permission_classes = [IsAdministrator]

    @extend_schema(
        request={"multipart/form-data": TechnicianImportSerializer},
        responses={201: ImportResultSerializer, 207: ImportResultSerializer},
    )
    def post(self, request):
        upload = request.FILES.get("file")
        if not upload or not upload.name.lower().endswith((".xlsx", ".xlsm")):
            return response.Response({"detail": "Adjunta un archivo Excel .xlsx o .xlsm."}, status=400)
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception:
            return response.Response({"detail": "No se pudo leer el archivo Excel."}, status=400)
        if not rows:
            return response.Response({"detail": "El archivo no contiene filas."}, status=400)
        headers = {str(value or "").strip().lower().replace(" ", "_"): index for index, value in enumerate(rows[0])}
        required = {"nombre", "codigo_trabajador", "dni"}
        if not required.issubset(headers):
            return response.Response({"detail": "Columnas requeridas: nombre, codigo_trabajador y dni. Opcionales: correo, especialidad, activo, contraseña_temporal."}, status=400)
        result = {"created": 0, "updated": 0, "errors": []}
        for number, row in enumerate(rows[1:], start=2):
            values = {key: (row[index] if index < len(row) else "") for key, index in headers.items()}
            try:
                full_name = str(values.get("nombre") or "").strip()
                worker_code = str(values.get("codigo_trabajador") or "").strip().upper()
                dni = "".join(ch for ch in str(values.get("dni") or "") if ch.isdigit())
                if not full_name or not worker_code or len(dni) != 8:
                    raise ValueError("nombre, codigo_trabajador y DNI de 8 dígitos son obligatorios")
                with transaction.atomic():
                    existing = AccountProfile.objects.filter(worker_code__iexact=worker_code).select_related("user").first()
                    if existing:
                        if existing.dni and existing.dni != dni:
                            raise ValueError("el código ya existe con otro DNI")
                        existing.dni = dni
                        existing.specialty = str(values.get("especialidad") or "").strip()
                        existing.position = str(values.get("cargo") or values.get("posicion") or "").strip()
                        existing.hourly_rate = values.get("tarifa_hora") or values.get("cuota_hora") or 0
                        existing.save(update_fields=("dni", "specialty", "position", "hourly_rate"))
                        result["updated"] += 1
                    else:
                        first_name, _, last_name = full_name.partition(" ")
                        password = str(values.get("contraseña_temporal") or "Importar2026!")
                        user = get_user_model().objects.create_user(username=worker_code.lower(), password=password, first_name=first_name, last_name=last_name, email=str(values.get("correo") or ""), is_active=True)
                        AccountProfile.objects.create(user=user, worker_code=worker_code, dni=dni, specialty=str(values.get("especialidad") or "").strip(), position=str(values.get("cargo") or values.get("posicion") or "").strip(), hourly_rate=values.get("tarifa_hora") or values.get("cuota_hora") or 0, role=AccountProfile.Role.TECHNICIAN, must_change_password=True)
                        result["created"] += 1
            except Exception as exc:
                result["errors"].append({"fila": number, "detalle": str(exc)})
        return response.Response(result, status=201 if not result["errors"] else 207)


class TechnicianDetailView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAdministrator]
    serializer_class = TechnicianSerializer
    queryset = get_user_model().objects.select_related('account_profile').filter(
        account_profile__role__in=[
            AccountProfile.Role.TECHNICIAN,
            AccountProfile.Role.ALMACENERO,
        ]
    )

    def get_object(self):
        return get_object_or_404(self.get_queryset(), account_profile__id=self.kwargs['pk'])


class TechnicianManualNotificationView(views.APIView):
    permission_classes = [IsAdministrator]

    @extend_schema(
        request=TechnicianNotificationSerializer,
        responses={201: DetailResponseSerializer, 400: DetailResponseSerializer},
    )
    def post(self, request, pk):
        technician = get_object_or_404(TechnicianDetailView.queryset, account_profile__id=pk)
        template = str(request.data.get('template') or 'CUSTOM').upper()
        delivery_channel = str(request.data.get('deliveryChannel') or 'SISTEMA').upper()
        if delivery_channel not in {'SISTEMA', 'CORREO'}:
            return response.Response({'detail': 'Selecciona Sistema o Correo como canal de envío.'}, status=status.HTTP_400_BAD_REQUEST)
        templates = {
            'REMINDER': ('Recordatorio de jornada', 'Recuerda revisar tu agenda, iniciar el temporizador al comenzar y actualizar el avance de cada OT.'),
            'TRACEABILITY': ('Actualiza la trazabilidad de tus OT', 'Registra el inicio, tiempo trabajado y avance de las órdenes asignadas para mantener la trazabilidad al día.'),
            'SCHEDULE': ('Cambio en tu programación', 'Revisa tu agenda semanal: se registró una actualización en la programación de tus órdenes de trabajo.'),
        }
        if template == 'CUSTOM':
            subject = str(request.data.get('subject') or '').strip()
            body = str(request.data.get('body') or '').strip()
            if not subject or not body:
                return response.Response({'detail': 'Ingresa asunto y mensaje para la notificación personalizada.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            if template not in templates:
                return response.Response({'detail': 'Plantilla no válida.'}, status=status.HTTP_400_BAD_REQUEST)
            subject, body = templates[template]
        notification = queue_notification(
            event='TECHNICIAN_MANUAL_NOTIFICATION', recipient=technician,
            subject=subject, body=body, discriminator=f'manual:{template}:{timezone.now().isoformat()}',
            delivery_channel=delivery_channel,
        )
        if not notification:
            return response.Response({'detail': 'El técnico necesita un correo activo para recibir avisos por correo.'}, status=status.HTTP_400_BAD_REQUEST)
        detail = 'Aviso publicado en la bandeja del técnico.' if delivery_channel == 'SISTEMA' else 'Correo programado para el técnico.'
        return response.Response({'detail': detail}, status=status.HTTP_201_CREATED)


__all__ = ["LoginView", "TokenRefreshView", "CurrentUserView", "ChangePasswordView", "UserListView", "TechnicianListCreateView", "TechnicianDetailView", "TechnicianManualNotificationView"]
