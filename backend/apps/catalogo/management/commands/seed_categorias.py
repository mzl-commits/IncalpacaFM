"""
Management command: seed_categorias

Lee las hojas 'Categorías' y 'Subcategorías' de la plantilla Excel y crea
(o completa) esos registros en la base de datos, usando get_or_create para
no duplicar nada que ya exista.

Uso:
    python manage.py seed_categorias importacion/Plantilla_importacion_materiales.xlsx
    python manage.py seed_categorias importacion/Plantilla_importacion_materiales.xlsx --almacen ALM-HERR
"""
import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.catalogo.models import Almacen, Categoria, Subcategoria


class Command(BaseCommand):
    help = "Crea/completa Categorías y Subcategorías desde las hojas de referencia del Excel."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta al .xlsx de la plantilla.")
        parser.add_argument(
            "--almacen", type=str, default=None,
            help="Código del almacén al que pertenecen las categorías. "
                 "Si no se indica y solo hay un almacén registrado, se usa ese.",
        )

    def handle(self, *args, **options):
        ruta = options["archivo"]
        codigo_almacen = options["almacen"]

        if codigo_almacen:
            try:
                almacen = Almacen.objects.get(codigo=codigo_almacen)
            except Almacen.DoesNotExist:
                raise CommandError(f"No existe un almacén con código '{codigo_almacen}'.")
        else:
            total = Almacen.objects.count()
            if total == 0:
                raise CommandError("No hay ningún almacén registrado. Crea uno primero desde la UI.")
            if total > 1:
                nombres = ", ".join(f"{a.codigo}" for a in Almacen.objects.all())
                raise CommandError(
                    f"Hay {total} almacenes ({nombres}). Especifica --almacen <codigo>."
                )
            almacen = Almacen.objects.first()

        self.stdout.write(f"Usando almacén: {almacen.codigo} - {almacen.nombre}")

        wb = openpyxl.load_workbook(ruta, data_only=True)

        # ── Categorías ──────────────────────────────────────────────────────
        ws_cat = wb["Categorías"]
        creadas_cat = 0
        for row in ws_cat.iter_rows(min_row=2, values_only=True):
            nombre, prefijo = (row + (None, None))[:2]
            if not nombre:
                continue
            obj, created = Categoria.objects.get_or_create(
                almacen=almacen,
                nombre=str(nombre).strip(),
                defaults={"prefijo": str(prefijo or "").strip()[:3]},
            )
            if created:
                creadas_cat += 1
                self.stdout.write(f"  + Categoría creada: {obj.nombre} ({obj.prefijo})")

        # ── Subcategorías ───────────────────────────────────────────────────
        ws_sub = wb["Subcategorías"]
        creadas_sub = 0
        no_encontradas = []
        for row in ws_sub.iter_rows(min_row=2, values_only=True):
            cat_nombre, sub_nombre = (row + (None, None))[:2]
            if not cat_nombre or not sub_nombre:
                continue
            try:
                categoria = Categoria.objects.get(almacen=almacen, nombre=str(cat_nombre).strip())
            except Categoria.DoesNotExist:
                no_encontradas.append((cat_nombre, sub_nombre))
                continue
            obj, created = Subcategoria.objects.get_or_create(
                categoria=categoria,
                nombre=str(sub_nombre).strip(),
            )
            if created:
                creadas_sub += 1
                self.stdout.write(f"  + Subcategoría creada: {cat_nombre} / {obj.nombre}")

        self.stdout.write(self.style.SUCCESS(
            f"\nListo. {creadas_cat} categoría(s) nueva(s), {creadas_sub} subcategoría(s) nueva(s)."
        ))
        if no_encontradas:
            self.stdout.write(self.style.WARNING(
                f"No se pudo crear subcategoría para {len(no_encontradas)} fila(s) "
                f"porque su categoría no existe: {no_encontradas}"
            ))
