"""
Management command: importar_inventario_consolidado

Importa el archivo Inventario_Consolidado.xlsx con sus dos hojas:
  - 'Materiales (No Retornables)' (Consumibles, stock inicial, empaque)
  - 'Herramientas y EPP' (Herramientas, EPP, piezas individuales, frecuencias)

Posteriormente:
  - Vincula plantillas de inspección SST.
  - Registra la inspección inicial masiva del 21/07/2026.
  - Regenera el calendario del plan anual desde el 22/07/2026.

Uso:
    python manage.py importar_inventario_consolidado
    python manage.py importar_inventario_consolidado --dry-run
    python manage.py importar_inventario_consolidado Inventario_Consolidado.xlsx
"""
import os
import unicodedata
import openpyxl
from datetime import date, datetime
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone

from apps.catalogo.models import (
    Almacen, Categoria, Subcategoria, Material, Pieza, TipoManejoStock
)
from apps.catalogo.services import crear_piezas_sueltas
from apps.inspeccion.models import (
    PlantillaCriterio, Inspeccion, ProgramacionInspeccion, RespuestaCriterio, PlanInspeccionAnual
)

User = get_user_model()


def _norm(val):
    if val is None:
        return ""
    return str(val).strip()


def _clave(val):
    s = _norm(val).casefold()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


# Prefijos por categoría para generación de códigos
PREFIJOS = {
    "acabados": "AC",
    "albanileria": "AL",
    "carpinteria": "C",
    "electricidad": "E",
    "equipo de proteccion personal": "EP",
    "epp": "EPP",
    "ferreteria": "F",
    "gasfiteria": "G",
    "herramientas": "H",
    "luminaria": "L",
    "redes y telecomunicaciones": "R",
    "proteccion para el rostro": "PR",
}

# Mapeo de Subcategorías a PlantillaCriterio
MAPEO_PLANTILLAS = {
    # Herramientas
    ("herramientas", "manual"): "Manual",
    ("herramientas", "manuales"): "Manual",
    ("herramientas", "electrica"): "Eléctrica",
    ("herramientas", "a bateria"): "Inalámbrica",
    ("herramientas", "inalambrica"): "Inalámbrica",
    ("herramientas", "complementos y repuestos"): "Manual",
    # Albañilería
    ("albanileria", "paletas y espatulas"): "Manual",
    ("albanileria", "pintura"): "Manual",
    # EPP
    ("equipo de proteccion personal", "proteccion corporal"): "Equipo de Protección Personal (EPP)",
    ("equipo de proteccion personal", "proteccion visual"): "Equipo de Protección Personal (EPP)",
    ("equipo de proteccion personal", "proteccion para el rostro"): "Equipo de Protección Personal (EPP)",
    ("equipo de proteccion personal", "proteccion contra caidas"): "Equipo de Protección Personal (EPP)",
    ("equipo de proteccion personal", "otros"): "Equipo de Protección Personal (EPP)",
    ("proteccion para el rostro", "caretas y visores"): "Equipo de Protección Personal (EPP)",
    ("proteccion para el rostro", "otros"): "Equipo de Protección Personal (EPP)",
}


class Command(BaseCommand):
    help = "Importa Inventario_Consolidado.xlsx y genera inspecciones y plan anual."

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            type=str,
            nargs="?",
            default="Inventario_Consolidado.xlsx",
            help="Ruta al archivo Inventario_Consolidado.xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula la importación sin guardar cambios en la base de datos.",
        )
        parser.add_argument(
            "--limpiar-todo",
            action="store_true",
            help="Limpia todos los materiales, piezas e inspecciones previas antes de importar el nuevo inventario.",
        )
        parser.add_argument(
            "--almacen",
            type=str,
            default="ALM-HERR",
            help="Código del almacén destino (por defecto ALM-HERR).",
        )

    def handle(self, *args, **options):
        archivo = options["archivo"]
        dry_run = options["dry_run"]
        codigo_almacen = options["almacen"]
        limpiar_todo = options["limpiar_todo"]

        if not os.path.exists(archivo):
            alt = os.path.join(os.path.dirname(__file__), "..", "..", "..", archivo)
            alt = os.path.abspath(alt)
            if os.path.exists(alt):
                archivo = alt
            else:
                alt2 = os.path.join(os.getcwd(), "backend", archivo)
                if os.path.exists(alt2):
                    archivo = alt2
                else:
                    raise CommandError(f"No se encontró el archivo: {archivo}")

        self.stdout.write(self.style.HTTP_INFO(f"==> Leyendo archivo: {archivo}"))
        wb = openpyxl.load_workbook(archivo, data_only=True)

        # 1. Resolver Almacén
        try:
            almacen = Almacen.objects.get(codigo=codigo_almacen)
        except Almacen.DoesNotExist:
            almacen = Almacen.objects.first()
            if not almacen:
                almacen = Almacen.objects.create(codigo=codigo_almacen, nombre="Almacén de Herramientas")

        self.stdout.write(f"Almacén destino: {almacen.codigo} - {almacen.nombre}")

        # Limpieza previa si se solicitó
        if limpiar_todo:
            if dry_run:
                self.stdout.write(self.style.NOTICE("  [DRY RUN] Se limpiarian todos los materiales, piezas e inspecciones previas."))
            else:
                self.stdout.write(self.style.WARNING("\n[!] Limpiando catalogo, piezas e inspecciones anteriores..."))
                RespuestaCriterio.objects.all().delete()
                ProgramacionInspeccion.objects.all().delete()
                Inspeccion.objects.all().delete()
                PlanInspeccionAnual.objects.all().delete()
                try:
                    from apps.inventario.models import Movimiento, SolicitudMovimiento
                    Movimiento.objects.all().delete()
                    SolicitudMovimiento.objects.all().delete()
                except Exception:
                    pass
                Pieza.objects.all().delete()
                Material.objects.all().delete()
                Subcategoria.objects.all().delete()
                Categoria.objects.all().delete()
                self.stdout.write(self.style.SUCCESS("  [OK] Base de datos limpia. Importando exclusivamente el nuevo inventario..."))

        # 2. Precargar tipos de manejo de stock
        tipos_manejo_map = {}
        for tm in TipoManejoStock.objects.all():
            tipos_manejo_map[_clave(tm.codigo)] = tm
            tipos_manejo_map[_clave(tm.nombre)] = tm
            nombre_sin_por = tm.nombre.lower()
            if nombre_sin_por.startswith("por "):
                tipos_manejo_map[_clave(nombre_sin_por[4:])] = tm

        alias_manejo = {
            "unidad": "unidad", "und": "unidad", "u": "unidad", "pza": "unidad", "pieza": "unidad",
            "caja": "cj", "cajas": "cj", "cj": "cj",
            "paquete": "paquete", "pqt": "paquete", "bolsa": "bolsa", "balde": "balde",
            "rollo": "rollo", "tubo": "tub", "galon": "galon", "bidon": "bidon",
            "par": "par", "set": "set", "kit": "kit", "millar": "millar", "docena": "docena",
        }
        for alias_k, target_code in alias_manejo.items():
            t_obj = tipos_manejo_map.get(_clave(target_code))
            if t_obj:
                tipos_manejo_map[_clave(alias_k)] = t_obj

        default_tm = TipoManejoStock.objects.filter(codigo="unidad").first() or TipoManejoStock.objects.first()

        # 3. Precargar plantillas de inspección
        plantillas_map = {}
        for p in PlantillaCriterio.objects.all():
            plantillas_map[_clave(p.nombre)] = p

        # Precargar categorías y subcategorías
        categorias_map = { _clave(c.nombre): c for c in Categoria.objects.filter(almacen=almacen) }
        subcategorias_map = { (s.categoria_id, _clave(s.nombre)): s for s in Subcategoria.objects.filter(categoria__almacen=almacen) }

        def obtener_o_crear_categoria(cat_nombre, requiere_insp=False):
            clave_c = _clave(cat_nombre)
            if clave_c in categorias_map:
                cat = categorias_map[clave_c]
                if requiere_insp and not cat.requiere_inspeccion and not dry_run:
                    cat.requiere_inspeccion = True
                    cat.save(update_fields=["requiere_inspeccion"])
                return cat
            base_prefijo = PREFIJOS.get(clave_c, cat_nombre[:3].upper())
            prefijo = base_prefijo[:3]
            if not dry_run:
                existentes_prefijos = set(Categoria.objects.filter(almacen=almacen).values_list("prefijo", flat=True))
                if prefijo in existentes_prefijos:
                    for i in range(1, 100):
                        cand = f"{base_prefijo[:2]}{i}"[:3]
                        if cand not in existentes_prefijos:
                            prefijo = cand
                            break
                cat = Categoria.objects.create(
                    almacen=almacen,
                    nombre=cat_nombre,
                    prefijo=prefijo,
                    requiere_inspeccion=requiere_insp,
                )
                categorias_map[clave_c] = cat
                return cat
            else:
                return Categoria(almacen=almacen, nombre=cat_nombre, prefijo=prefijo, requiere_inspeccion=requiere_insp)

        def obtener_o_crear_subcategoria(categoria, sub_nombre):
            if dry_run and not getattr(categoria, "id", None):
                return Subcategoria(categoria=categoria, nombre=sub_nombre)
            clave_s = (categoria.id, _clave(sub_nombre))
            if clave_s in subcategorias_map:
                sub_existente = subcategorias_map[clave_s]
                # Si no tiene plantilla y hay mapeo, asignarla
                if not sub_existente.plantilla_inspeccion_id and not dry_run:
                    clave_cat = _clave(categoria.nombre)
                    clave_sub = _clave(sub_nombre)
                    plantilla_nombre_target = MAPEO_PLANTILLAS.get((clave_cat, clave_sub))
                    if plantilla_nombre_target and _clave(plantilla_nombre_target) in plantillas_map:
                        sub_existente.plantilla_inspeccion = plantillas_map[_clave(plantilla_nombre_target)]
                        sub_existente.save(update_fields=["plantilla_inspeccion"])
                return sub_existente
            
            # Buscar si tiene plantilla asignada según mapeo
            plantilla_obj = None
            clave_cat = _clave(categoria.nombre)
            clave_sub = _clave(sub_nombre)
            plantilla_nombre_target = MAPEO_PLANTILLAS.get((clave_cat, clave_sub))
            if plantilla_nombre_target:
                plantilla_obj = plantillas_map.get(_clave(plantilla_nombre_target))

            if not dry_run:
                sub = Subcategoria.objects.create(
                    categoria=categoria,
                    nombre=sub_nombre,
                    plantilla_inspeccion=plantilla_obj,
                )
                subcategorias_map[clave_s] = sub
                return sub
            else:
                return Subcategoria(categoria=categoria, nombre=sub_nombre, plantilla_inspeccion=plantilla_obj)

        with transaction.atomic():
            sid = transaction.savepoint()
            total_no_retornables = 0
            total_herramientas = 0
            total_piezas_creadas = 0

            # =========================================================================
            # HOJA 1: Materiales (No Retornables)
            # =========================================================================
            if "Materiales (No Retornables)" in wb.sheetnames:
                ws_nr = wb["Materiales (No Retornables)"]
                self.stdout.write(f"\n[>>>] Procesando hoja: 'Materiales (No Retornables)' ({ws_nr.max_row - 1} filas)...")
                
                header = [_norm(c.value) for c in ws_nr[1]]
                idx = { h: i for i, h in enumerate(header) if h }

                for row in ws_nr.iter_rows(min_row=2, values_only=True):
                    if not row or all(v is None for v in row):
                        continue
                    cat_nom = _norm(row[idx.get("Categoría*", 0)])
                    sub_nom = _norm(row[idx.get("Subcategoría*", 1)])
                    nombre = _norm(row[idx.get("Nombre*", 2)])
                    if not nombre or not cat_nom or not sub_nom:
                        continue

                    marca = _norm(row[idx.get("Marca", 3)])
                    modelo = _norm(row[idx.get("Modelo", 4)])
                    medida = _norm(row[idx.get("Medida", 5)])
                    precio = row[idx.get("Precio (S/)", 6)] or None
                    ubicacion = _norm(row[idx.get("Ubicación física*", 7)]) or "Almacén General"
                    unidad_manejo_raw = _norm(row[idx.get("Unidad de manejo de stock*", 8)])
                    unidades_empaque = row[idx.get("Unidades por empaque", 9)]
                    stock_inicial = row[idx.get("Stock total inicial (en unidades)*", 10)] or 0
                    stock_min = row[idx.get("Stock mínimo de alerta", 11)] or 0
                    notas = _norm(row[idx.get("Notas / observaciones", 13)])

                    categoria = obtener_o_crear_categoria(cat_nom, requiere_insp=False)
                    subcategoria = obtener_o_crear_subcategoria(categoria, sub_nom)

                    tm_obj = tipos_manejo_map.get(_clave(unidad_manejo_raw), default_tm)

                    try:
                        stock_ini_int = int(stock_inicial)
                    except Exception:
                        stock_ini_int = 0
                    try:
                        stock_min_int = int(stock_min)
                    except Exception:
                        stock_min_int = 0
                    try:
                        unidades_emp_int = int(unidades_empaque) if unidades_empaque else None
                    except Exception:
                        unidades_emp_int = None

                    if not dry_run:
                        mat, created = Material.objects.get_or_create(
                            almacen=almacen,
                            subcategoria=subcategoria,
                            nombre=nombre,
                            medida=medida,
                            marca=marca,
                            defaults=dict(
                                modelo=modelo,
                                precio=precio if isinstance(precio, (int, float)) else None,
                                ubicacion_fisica=ubicacion,
                                tipo_control="no_retornable",
                                control_individual=False,
                                unidad_manejo=tm_obj,
                                unidades_por_caja=unidades_emp_int,
                                cantidad_total=stock_ini_int,
                                stock_minimo=stock_min_int,
                            ),
                        )
                        if not created:
                            mat.modelo = modelo
                            mat.ubicacion_fisica = ubicacion
                            mat.unidad_manejo = tm_obj
                            mat.unidades_por_caja = unidades_emp_int
                            mat.cantidad_total = stock_ini_int
                            mat.stock_minimo = stock_min_int
                            mat.save()

                    total_no_retornables += 1

            # =========================================================================
            # HOJA 2: Herramientas y EPP
            # =========================================================================
            if "Herramientas y EPP" in wb.sheetnames:
                ws_h = wb["Herramientas y EPP"]
                self.stdout.write(f"\n[>>>] Procesando hoja: 'Herramientas y EPP' ({ws_h.max_row - 1} filas)...")
                
                header_h = [_norm(c.value) for c in ws_h[1]]
                idx_h = { h: i for i, h in enumerate(header_h) if h }

                for row in ws_h.iter_rows(min_row=2, values_only=True):
                    if not row or all(v is None for v in row):
                        continue
                    cat_nom = _norm(row[idx_h.get("Categoría*", 0)])
                    sub_nom = _norm(row[idx_h.get("Subcategoría*", 1)])
                    subtipo = _norm(row[idx_h.get("Subtipo", 2)])
                    nombre = _norm(row[idx_h.get("Nombre*", 3)])
                    if not nombre or not cat_nom or not sub_nom:
                        continue

                    marca = _norm(row[idx_h.get("Marca", 4)])
                    modelo = _norm(row[idx_h.get("Modelo", 5)])
                    medida = _norm(row[idx_h.get("Medida", 6)])
                    precio = row[idx_h.get("Precio (S/)", 7)] or None
                    frec_val = row[idx_h.get("Frecuencia inspección - valor*", 8)] or 3
                    frec_uni = _norm(row[idx_h.get("Frecuencia inspección - unidad*", 9)]).lower() or "meses"
                    if frec_uni not in ("dias", "meses"):
                        frec_uni = "meses"
                    ubicacion = _norm(row[idx_h.get("Ubicación física*", 10)]) or "Almacén General"
                    cant_piezas_raw = row[idx_h.get("Cantidad de piezas / estuches a crear*", 11)] or 1
                    notas = _norm(row[idx_h.get("Notas / observaciones", 13)])

                    try:
                        cant_piezas = int(cant_piezas_raw)
                    except Exception:
                        cant_piezas = 1

                    try:
                        frec_val_int = int(frec_val)
                    except Exception:
                        frec_val_int = 3

                    categoria = obtener_o_crear_categoria(cat_nom, requiere_insp=True)
                    subcategoria = obtener_o_crear_subcategoria(categoria, sub_nom)

                    if not dry_run:
                        mat, created = Material.objects.get_or_create(
                            almacen=almacen,
                            subcategoria=subcategoria,
                            nombre=nombre,
                            medida=medida,
                            marca=marca,
                            defaults=dict(
                                modelo=modelo,
                                precio=precio if isinstance(precio, (int, float)) else None,
                                ubicacion_fisica=ubicacion,
                                tipo_control="retornable",
                                control_individual=True,
                                periodicidad_valor=frec_val_int,
                                periodicidad_unidad=frec_uni,
                            ),
                        )
                        if not created:
                            mat.modelo = modelo
                            mat.ubicacion_fisica = ubicacion
                            mat.periodicidad_valor = frec_val_int
                            mat.periodicidad_unidad = frec_uni
                            mat.save()

                        # Crear piezas individuales sueltas si no existen
                        piezas_existentes = mat.piezas.exclude(estado="Baja").count()
                        if piezas_existentes < cant_piezas:
                            creadas_piezas = crear_piezas_sueltas(mat, cant_piezas - piezas_existentes)
                            total_piezas_creadas += len(creadas_piezas)
                        else:
                            total_piezas_creadas += piezas_existentes
                    else:
                        total_piezas_creadas += cant_piezas

                    total_herramientas += 1

            if dry_run:
                transaction.savepoint_rollback(sid)
                self.stdout.write(self.style.NOTICE(
                    f"\n[DRY RUN] Validación exitosa:\n"
                    f"  - No Retornables listos para importar: {total_no_retornables}\n"
                    f"  - Herramientas y EPP listos: {total_herramientas}\n"
                    f"  - Piezas individuales a generar: {total_piezas_creadas}"
                ))
                return

            transaction.savepoint_commit(sid)
            self.stdout.write(self.style.SUCCESS(
                f"\n[OK] Importacion de materiales completada:\n"
                f"  - No Retornables importados: {total_no_retornables}\n"
                f"  - Herramientas y EPP importados: {total_herramientas}\n"
                f"  - Piezas creadas/activas: {total_piezas_creadas}"
            ))

        # =========================================================================
        # PASO 2: Vincular Plantillas de Inspección
        # =========================================================================
        self.stdout.write("\n[>>>] Vinculando plantillas de inspeccion SST a subcategorias...")
        call_command("vincular_plantillas", stdout=self.stdout, stderr=self.stderr)

        # =========================================================================
        # PASO 3: Inicializar Inspecciones Masivas del 21/07/2026
        # =========================================================================
        self.stdout.write("\n[>>>] Inicializando inspecciones iniciales del 21/07/2026...")
        call_command("inicializar_inspecciones_julio", stdout=self.stdout, stderr=self.stderr)

        # =========================================================================
        # PASO 4: Regenerar Plan Anual 2026 desde el 22/07/2026
        # =========================================================================
        self.stdout.write("\n[>>>] Regenerando plan anual de inspecciones 2026 (max 5/dia)...")
        call_command(
            "reiniciar_plan_desde_julio",
            fecha_inicio="2026-07-22",
            max_por_dia=5,
            stdout=self.stdout,
            stderr=self.stderr,
        )

        self.stdout.write(self.style.SUCCESS("\n========================================================"))
        self.stdout.write(self.style.SUCCESS("[OK] PROCESO COMPLETO FINALIZADO CON EXITO"))
        self.stdout.write(self.style.SUCCESS("========================================================"))
