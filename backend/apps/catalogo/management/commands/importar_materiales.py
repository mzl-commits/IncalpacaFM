"""
Management command: importar_materiales

Lee la plantilla Excel (hojas 'Materiales' y 'Piezas_Estuche') y carga los
datos en la base de datos usando el ORM de Django (no INSERTs manuales), para
poder correr este mismo script primero en el entorno de pruebas y luego,
sin cambios, contra la base de datos real en producción.

Uso:
    python manage.py importar_materiales ruta/al/archivo.xlsx
    python manage.py importar_materiales ruta/al/archivo.xlsx --dry-run
    python manage.py importar_materiales ruta/al/archivo.xlsx --fotos-dir ruta/a/fotos/
    python manage.py importar_materiales ruta/al/archivo.xlsx --dry-run --permitir-sin-foto

Ubicación sugerida:
    backend/apps/catalogo/management/commands/importar_materiales.py
    (crear también, si no existen, los __init__.py vacíos en
     apps/catalogo/management/ y apps/catalogo/management/commands/)

NOTA IMPORTANTE (fix aplicado):
    La comparación de Categoría/Subcategoría ya NO se hace con __iexact en la
    base de datos. Motivo: en SQLite, __iexact solo pliega mayúsculas/minúsculas
    de forma confiable para letras sin tilde (a-z). Con vocales acentuadas
    (Á, É, Í, Ó, Ú, Ñ) el case-insensitive de SQLite puede fallar y reportar
    "no existe" aunque el texto sea, para el ojo humano, exactamente el mismo
    (ej. 'protección' vs 'PROTECCIÓN CONTRA CAÍDAS' con distinta mayúscula en
    una vocal tildeada). Para evitarlo, se cargan todas las categorías y
    subcategorías una sola vez en memoria y se comparan normalizadas en
    Python con .strip().casefold(), que sí es correcto con acentos.
"""
import os
import unicodedata
import openpyxl
from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalogo.models import (
    Categoria, Subcategoria, Material, Pieza, TipoManejoStock
)
from apps.catalogo.services import crear_piezas_sueltas, crear_estuche_con_piezas, ajustar_stock


COLS = {
    "categoria": "Categoría*",
    "subcategoria": "Subcategoría*",
    "nombre": "Nombre*",
    "marca": "Marca",
    "modelo": "Modelo",
    "medida": "Medida",
    "precio": "Precio (S/)",
    "periodicidad_valor": "Frecuencia inspección - valor*",
    "periodicidad_unidad": "Frecuencia inspección - unidad*",
    "ubicacion_fisica": "Ubicación física*",
    "tipo_control": "Tipo de control*",
    "control_individual": "Control por pieza individual*",
    "unidad_manejo": "Unidad de manejo (si no es individual)",
    "unidades_por_caja": "Unidades por empaque",
    "cantidad_piezas": "Cantidad de piezas sueltas a crear",
    "foto": "Foto representativa (ruta archivo)*",
    "notas": "Notas / observaciones",
}

REQUIRED_KEYS = [
    "categoria", "subcategoria", "nombre", "periodicidad_valor",
    "periodicidad_unidad", "ubicacion_fisica", "tipo_control",
    "control_individual", "foto",
]


def _norm(value):
    if value is None:
        return ""
    return str(value).strip()


def _clave(value):
    """Normaliza para comparar sin problemas de mayúsculas NI tildes."""
    s = _norm(value).casefold()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s


def _si_no_a_bool(value):
    v = _norm(value).lower()
    if v in ("sí", "si", "true", "1", "x", "yes"):
        return True
    if v in ("no", "false", "0", ""):
        return False
    raise ValueError(f"Valor de 'Control por pieza individual' no reconocido: {value!r}")


class Command(BaseCommand):
    help = "Importa materiales (y piezas) desde la plantilla Excel usando el ORM de Django."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta al .xlsx de la plantilla.")
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Valida todo el archivo pero no escribe nada en la base de datos.",
        )
        parser.add_argument(
            "--fotos-dir", type=str, default=None,
            help="Carpeta donde buscar las fotos referenciadas en la columna 'Foto representativa' "
                 "si la ruta del Excel no es absoluta ni existe tal cual.",
        )
        parser.add_argument(
            "--permitir-sin-foto", action="store_true",
            help="Permite importar materiales sin foto representativa (para completarla después "
                 "desde la UI). Sin este flag, la foto es obligatoria y una fila sin foto es un error.",
        )

    def handle(self, *args, **options):
        ruta = options["archivo"]
        dry_run = options["dry_run"]
        fotos_dir = options["fotos_dir"]
        permitir_sin_foto = options["permitir_sin_foto"]

        if not os.path.exists(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        wb = openpyxl.load_workbook(ruta, data_only=True)
        if "Materiales" not in wb.sheetnames:
            raise CommandError("El archivo no tiene una hoja llamada 'Materiales'.")

        ws = wb["Materiales"]
        header = [_norm(c.value) for c in ws[1]]
        col_idx = {}
        for key, label in COLS.items():
            if label in header:
                col_idx[key] = header.index(label)
            elif key in REQUIRED_KEYS:
                raise CommandError(f"Falta la columna obligatoria '{label}' en la hoja Materiales.")

        # --- Precarga de categorías y subcategorías, normalizadas en Python ---
        # (evita el bug de __iexact + tildes en SQLite, ver nota al inicio del archivo)
        categorias_por_clave = {}
        for cat in Categoria.objects.all():
            categorias_por_clave[_clave(cat.nombre)] = cat

        subcategorias_por_clave = {}
        subcategorias_por_categoria_id = {}
        for sub in Subcategoria.objects.select_related("categoria__almacen").all():
            subcategorias_por_clave[(sub.categoria_id, _clave(sub.nombre))] = sub
            subcategorias_por_categoria_id.setdefault(sub.categoria_id, []).append(sub.nombre)

        # --- Precarga de tipos de manejo de stock ---
        # Asegurar tipo de manejo 'balde' si no existe
        TipoManejoStock.objects.get_or_create(
            codigo="balde",
            defaults={"nombre": "Por Balde", "requiere_multiplicador": True, "orden": 31},
        )
        tipos_manejo_map = {}
        for tm in TipoManejoStock.objects.all():
            tipos_manejo_map[_clave(tm.codigo)] = tm
            tipos_manejo_map[_clave(tm.nombre)] = tm
            # También mapear sin prefijo "por " si aplica
            nombre_sin_por = tm.nombre.lower()
            if nombre_sin_por.startswith("por "):
                tipos_manejo_map[_clave(nombre_sin_por[4:])] = tm

        # Alias comunes
        alias_manejo = {
            "caja": "cj",
            "cajas": "cj",
            "paquete": "paquete",
            "paquetes": "paquete",
            "pqt": "paquete",
            "bolsa": "bolsa",
            "bolsas": "bolsa",
            "balde": "balde",
            "baldes": "balde",
            "unidad": "unidad",
            "unidades": "unidad",
            "und": "unidad",
            "u": "unidad",
            "rollo": "rollo",
            "rollos": "rollo",
            "tubo": "tub",
            "tubos": "tub",
            "blister": "blister",
            "kit": "kit",
            "juego": "kit",
            "docena": "docena",
            "millar": "millar",
        }
        for alias_k, target_code in alias_manejo.items():
            target_obj = tipos_manejo_map.get(_clave(target_code))
            if target_obj:
                tipos_manejo_map[_clave(alias_k)] = target_obj

        default_tm = TipoManejoStock.objects.filter(codigo="unidad").first() or TipoManejoStock.objects.first()

        ws_est = wb["Piezas_Estuche"] if "Piezas_Estuche" in wb.sheetnames else None
        piezas_estuche_por_material = {}
        if ws_est is not None:
            for row in ws_est.iter_rows(min_row=2, values_only=True):
                if row is None or all(v is None for v in row):
                    continue
                mat_nombre = _norm(row[0])
                pieza_nombre = _norm(row[1])
                medida = _norm(row[2]) if len(row) > 2 else ""
                cantidad = int(row[3]) if len(row) > 3 and row[3] else 0
                if not mat_nombre or not pieza_nombre or cantidad <= 0:
                    continue
                piezas_estuche_por_material.setdefault(mat_nombre, []).append(
                    {"nombre": pieza_nombre, "medida": medida, "cantidad": cantidad}
                )

        creados, saltados, errores = 0, 0, []

        with transaction.atomic():
            sid = transaction.savepoint()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None or all(v is None for v in row):
                    continue
                nombre = _norm(row[col_idx["nombre"]])
                if not nombre:
                    saltados += 1
                    continue

                try:
                    cat_nombre = _norm(row[col_idx["categoria"]])
                    sub_nombre = _norm(row[col_idx["subcategoria"]])
                    if not cat_nombre or not sub_nombre:
                        raise ValueError("Categoría y Subcategoría son obligatorias.")

                    categoria = categorias_por_clave.get(_clave(cat_nombre))
                    if categoria is None:
                        disponibles = ", ".join(sorted(c.nombre for c in categorias_por_clave.values())) or "(ninguna en el sistema)"
                        raise ValueError(
                            f"Categoría no encontrada: '{cat_nombre}'. Categorías existentes: {disponibles}"
                        )

                    subcategoria = subcategorias_por_clave.get((categoria.id, _clave(sub_nombre)))
                    if subcategoria is None:
                        disponibles = ", ".join(sorted(subcategorias_por_categoria_id.get(categoria.id, []))) or "(ninguna)"
                        raise ValueError(
                            f"Subcategoría '{sub_nombre}' no existe para '{cat_nombre}'. "
                            f"Subcategorías existentes en '{cat_nombre}': {disponibles}"
                        )

                    tipo_control = _norm(row[col_idx["tipo_control"]]).lower()
                    if tipo_control not in ("retornable", "no_retornable"):
                        raise ValueError(f"Tipo de control inválido: '{tipo_control}'.")

                    control_individual = _si_no_a_bool(row[col_idx["control_individual"]])

                    periodicidad_unidad = _norm(row[col_idx["periodicidad_unidad"]]).lower()
                    if periodicidad_unidad not in ("dias", "meses"):
                        raise ValueError(f"Unidad de periodicidad inválida: '{periodicidad_unidad}'.")
                    periodicidad_valor = int(row[col_idx["periodicidad_valor"]])

                    ubicacion = _norm(row[col_idx["ubicacion_fisica"]])
                    if not ubicacion:
                        raise ValueError("Ubicación física es obligatoria.")

                    foto_ref = _norm(row[col_idx["foto"]])
                    if not foto_ref and not permitir_sin_foto:
                        raise ValueError(
                            "Foto representativa es obligatoria (usa --permitir-sin-foto para "
                            "importar sin ella y completarla después)."
                        )

                    almacen_obj = subcategoria.categoria.almacen

                    material_kwargs = dict(
                        subcategoria=subcategoria,
                        almacen=almacen_obj,
                        nombre=nombre,
                        marca=_norm(row[col_idx["marca"]]),
                        modelo=_norm(row[col_idx["modelo"]]),
                        medida=_norm(row[col_idx["medida"]]),
                        precio=row[col_idx["precio"]] or None,
                        ubicacion_fisica=ubicacion,
                        tipo_control=tipo_control,
                        control_individual=control_individual,
                        periodicidad_valor=periodicidad_valor,
                        periodicidad_unidad=periodicidad_unidad,
                    )
                    if not control_individual:
                        val_umanejo = _norm(row[col_idx["unidad_manejo"]])
                        if val_umanejo:
                            tm_obj = tipos_manejo_map.get(_clave(val_umanejo))
                            if not tm_obj:
                                raise ValueError(f"Tipo de manejo de stock no reconocido: '{val_umanejo}'.")
                        else:
                            tm_obj = default_tm

                        material_kwargs["unidad_manejo"] = tm_obj
                        if tm_obj.requiere_multiplicador:
                            upc = row[col_idx["unidades_por_caja"]]
                            if not upc:
                                raise ValueError(
                                    f"Unidades por empaque es obligatorio cuando la unidad de manejo es '{tm_obj.nombre}'."
                                )
                            material_kwargs["unidades_por_caja"] = int(upc)

                    if dry_run:
                        creados += 1
                        continue

                    material = Material.objects.create(**material_kwargs)

                    if foto_ref:
                        foto_path = _resolver_foto(foto_ref, fotos_dir)
                        if foto_path:
                            with open(foto_path, "rb") as f:
                                material.foto.save(os.path.basename(foto_path), File(f), save=True)
                        else:
                            self.stdout.write(self.style.WARNING(
                                f"  Fila {i}: no se encontró la foto '{foto_ref}', se importó sin foto."
                            ))
                    elif permitir_sin_foto:
                        self.stdout.write(self.style.WARNING(
                            f"  Fila {i}: sin foto (--permitir-sin-foto activo), completar después desde la UI."
                        ))

                    if not control_individual:
                        cantidad_inicial = row[col_idx["cantidad_piezas"]]
                        if cantidad_inicial:
                            ajustar_stock(material, int(cantidad_inicial))
                    else:
                        if nombre in piezas_estuche_por_material:
                            spec = []
                            for item in piezas_estuche_por_material[nombre]:
                                mat_hija, _ = Material.objects.get_or_create(
                                    subcategoria=subcategoria,
                                    nombre=item["nombre"],
                                    medida=item["medida"],
                                    es_componente=True,
                                    defaults=dict(
                                        almacen=almacen_obj,
                                        tipo_control=tipo_control,
                                        control_individual=True,
                                    ),
                                )
                                spec.append({"material": mat_hija, "cantidad": item["cantidad"]})
                            crear_estuche_con_piezas(material, spec, num_estuches=1)
                        else:
                            cantidad = row[col_idx["cantidad_piezas"]]
                            if cantidad:
                                crear_piezas_sueltas(material, int(cantidad))

                    creados += 1

                except Exception as e:
                    errores.append(f"Fila {i} ('{nombre}'): {e}")

            if errores:
                transaction.savepoint_rollback(sid)
            elif dry_run:
                transaction.savepoint_rollback(sid)
            else:
                transaction.savepoint_commit(sid)

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.NOTICE(f"[DRY RUN] Filas válidas: {creados} | Filas vacías: {saltados}"))
        else:
            self.stdout.write(self.style.SUCCESS(f"Materiales creados: {creados} | Filas vacías: {saltados}"))

        if errores:
            self.stdout.write(self.style.ERROR(f"\nSe encontraron {len(errores)} error(es). No se guardó nada:"))
            for err in errores:
                self.stdout.write(self.style.ERROR(f"  - {err}"))
            raise CommandError("Importación abortada por errores. Corrige el Excel y vuelve a intentar.")


def _resolver_foto(foto_ref, fotos_dir):
    """Intenta encontrar el archivo de foto: ruta tal cual, o dentro de --fotos-dir."""
    if os.path.isabs(foto_ref) and os.path.exists(foto_ref):
        return foto_ref
    if os.path.exists(foto_ref):
        return foto_ref
    if fotos_dir:
        candidato = os.path.join(fotos_dir, foto_ref)
        if os.path.exists(candidato):
            return candidato
        candidato2 = os.path.join(fotos_dir, os.path.basename(foto_ref))
        if os.path.exists(candidato2):
            return candidato2
    return None